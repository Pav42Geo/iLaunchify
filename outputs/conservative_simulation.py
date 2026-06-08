"""
iLaunchify financial simulation — Conservative vs Base side-by-side.

Halves the four most-optimistic assumptions:
  1. Activation rate per tier
  2. Orders per active creator per month
  3. Bulk runs per active creator per month
  4. Velocity distribution shifted downward (creators stuck at Tier 1 longer)

Everything else (pricing, partner mix, order economics, costs) stays identical
so the only difference is the demand-side assumptions.

This is the realistic year-one model. The Base scenario is "if everything
works as we hope after 18-24 months and we've found product-market fit."
"""

import sys
from copy import deepcopy

sys.path.insert(0, "./")

# Import the Base assumptions + core math from financial_simulation
import financial_simulation as base

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ─────────────────────────────────────────────────────────────────────
# CONSERVATIVE OVERRIDES — halve the optimistic levers
# ─────────────────────────────────────────────────────────────────────

CONSERVATIVE_ACTIVATION = {
    "maker":   0.15,    # was 0.30 → half of registered Makers actually use it
    "builder": 0.35,    # was 0.70 → paid subscribers still churn
    "agency":  0.45,    # was 0.90 → some Agency seats sit dormant
}

CONSERVATIVE_ORDERS_PER_ACTIVE = {
    "maker":   2,       # was 5 → real Maker median is very low
    "builder": 40,      # was 80 → realistic Shopify creator volume
    "agency":  200,     # was 400 → established but not whale
}

CONSERVATIVE_BULK_RUNS_PER_MONTH = {
    "maker":   0.05,    # was 0.10 → 1 every 20 months
    "builder": 0.25,    # was 0.50 → 1 per quarter
    "agency":  1.0,     # was 2.0 → 1 bulk run per month
}

# Velocity distribution shifted down — most creators stuck at Tier 1 longer
CONSERVATIVE_VELOCITY_DISTRIBUTION = {
    "maker":   {1: 0.95, 2: 0.04, 3: 0.01, 4: 0.00, 5: 0.00},  # was 90/8/2/0/0
    "builder": {1: 0.70, 2: 0.20, 3: 0.07, 4: 0.02, 5: 0.01},  # was 55/25/12/6/2
    "agency":  {1: 0.50, 2: 0.30, 3: 0.15, 4: 0.04, 5: 0.01},  # was 30/30/20/12/8
}

# ─────────────────────────────────────────────────────────────────────
# Re-run the simulator with patched assumptions for Conservative
# ─────────────────────────────────────────────────────────────────────

def run_with_overrides(creator_count, activation, orders, bulk_runs, velocity_dist):
    """Temporarily patch the financial_simulation module's assumptions and run simulate()."""
    saved = (
        deepcopy(base.CREATOR_ACTIVATION),
        deepcopy(base.ORDERS_PER_ACTIVE_CREATOR_PER_MONTH),
        deepcopy(base.BULK_RUNS_PER_MONTH),
        deepcopy(base.VELOCITY_DISTRIBUTION),
    )
    base.CREATOR_ACTIVATION = activation
    base.ORDERS_PER_ACTIVE_CREATOR_PER_MONTH = orders
    base.BULK_RUNS_PER_MONTH = bulk_runs
    base.VELOCITY_DISTRIBUTION = velocity_dist
    try:
        result = base.simulate(creator_count)
    finally:
        (base.CREATOR_ACTIVATION,
         base.ORDERS_PER_ACTIVE_CREATOR_PER_MONTH,
         base.BULK_RUNS_PER_MONTH,
         base.VELOCITY_DISTRIBUTION) = saved
    return result


SCALES = [100, 1_000, 10_000, 30_000]

base_results = {}
conservative_results = {}
for n in SCALES:
    base_results[n] = base.simulate(n)
    conservative_results[n] = run_with_overrides(
        n,
        CONSERVATIVE_ACTIVATION,
        CONSERVATIVE_ORDERS_PER_ACTIVE,
        CONSERVATIVE_BULK_RUNS_PER_MONTH,
        CONSERVATIVE_VELOCITY_DISTRIBUTION,
    )

# ─────────────────────────────────────────────────────────────────────
# Console print
# ─────────────────────────────────────────────────────────────────────

print("\n" + "═" * 100)
print("Base (optimistic) vs Conservative (halved demand levers)")
print("═" * 100)
print(f"\n{'Scale':<14} {'Scenario':<14} {'Ann revenue':>16} {'Op profit':>16} {'Profit/creator':>16}")
for n in SCALES:
    for label, r in [("Base", base_results[n]), ("Conservative", conservative_results[n])]:
        per_creator = r.annualized_profit / n if n else 0
        print(f"{n:>6,} creators  {label:<14} ${r.annualized_revenue:>14,.0f} ${r.annualized_profit:>14,.0f} ${per_creator:>14,.0f}")

print(f"\n{'Conservative as % of Base ARR':<40}")
for n in SCALES:
    pct = conservative_results[n].annualized_revenue / base_results[n].annualized_revenue * 100
    print(f"{n:>6,} creators → Conservative is {pct:.1f}% of Base ARR")

# ─────────────────────────────────────────────────────────────────────
# XLSX output
# ─────────────────────────────────────────────────────────────────────

wb = Workbook()
PINK = "FF2E63"
INK_900 = "111111"
CREAM = "F3EFE8"
LIGHT_PINK = "FFE0EB"
LIGHT_NEON = "EFFFCC"
LIGHT_GREY = "F5F5F5"

hdr_font = Font(name="Inter", size=12, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=INK_900)
section_font = Font(name="Inter", size=14, bold=True, color=INK_900)
section_fill = PatternFill("solid", fgColor=CREAM)
money_format = '"$"#,##0'

# ─── Sheet 1: Side-by-side comparison ───
ws = wb.active
ws.title = "Side-by-side"

ws["A1"] = "Conservative vs Base — year-one realistic vs PMF target"
ws["A1"].font = Font(name="Inter", size=18, bold=True, color=PINK)
ws.merge_cells("A1:J1")

ws["A2"] = "Same pricing model (V1.5 locked). Conservative halves activation + order velocity + bulk frequency + shifts velocity-tier distribution downward."
ws["A2"].font = Font(italic=True, size=10, color="666666")
ws.merge_cells("A2:J2")

# Header row
ws.cell(row=4, column=1, value="Metric").font = hdr_font
ws.cell(row=4, column=1).fill = hdr_fill
col = 2
for n in SCALES:
    for label in ("Conservative", "Base"):
        c = ws.cell(row=4, column=col, value=f"{label} · {n:,}")
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        col += 1

def write_row(label, key, fmt=money_format, bold=False):
    idx = ws.max_row + 1
    ws.cell(row=idx, column=1, value=label).font = Font(bold=bold)
    col = 2
    for n in SCALES:
        for source in (conservative_results, base_results):
            cell = ws.cell(row=idx, column=col, value=getattr(source[n], key))
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
            if source is conservative_results:
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            else:
                cell.fill = PatternFill("solid", fgColor=LIGHT_NEON)
            col += 1
    return idx

def section(label):
    idx = ws.max_row + 2
    c = ws.cell(row=idx, column=1, value=label)
    c.font = section_font
    c.fill = section_fill
    return idx

section("USER COUNTS")
write_row("Partners", "partner_count", "#,##0")
write_row("Active Makers", "active_makers", "#,##0")
write_row("Active Builders", "active_builders", "#,##0")
write_row("Active Agencies", "active_agencies", "#,##0")

section("MONTHLY REVENUE STREAMS")
write_row("Subscription revenue", "subscription_revenue_monthly")
write_row("On-demand units", "on_demand_units_monthly", "#,##0")
write_row("On-demand gross retail", "on_demand_gross_revenue_monthly")
write_row("Creator-side on-demand fee", "on_demand_platform_fee_creator_side")
write_row("Partner-side on-demand fee", "on_demand_platform_fee_partner_side")
write_row("Bulk units", "bulk_units_monthly", "#,##0")
write_row("Bulk partner cost (gross)", "bulk_gross_revenue_monthly")
write_row("Creator-side bulk fee", "bulk_platform_fee_creator_side")
write_row("Partner-side bulk fee", "bulk_platform_fee_partner_side")

section("MONTHLY COSTS")
write_row("Stripe Connect", "stripe_costs_monthly")
write_row("AI parser (Anthropic)", "ai_parser_costs_monthly")
write_row("Asset library", "asset_library_costs_monthly")
write_row("Fixed opex (eng+infra)", "fixed_opex_monthly")

section("BOTTOM LINE")
write_row("Total platform take (monthly)", "total_platform_take_monthly", bold=True)
write_row("Operating profit (monthly)", "operating_profit_monthly", bold=True)
write_row("Annualized revenue", "annualized_revenue", bold=True)
write_row("Annualized op profit", "annualized_profit", bold=True)

# Highlight last 2 rows extra
for r_idx in (ws.max_row - 1, ws.max_row):
    for col_idx in range(1, 10):
        cell = ws.cell(row=r_idx, column=col_idx)
        cell.font = Font(bold=True, size=12)

# Column widths
ws.column_dimensions["A"].width = 32
for col_letter in "BCDEFGHIJ":
    ws.column_dimensions[col_letter].width = 16

# ─── Sheet 2: Headline summary ───
ws2 = wb.create_sheet("Summary")

ws2["A1"] = "Headline numbers — Conservative vs Base"
ws2["A1"].font = Font(name="Inter", size=18, bold=True, color=PINK)
ws2.merge_cells("A1:E1")

ws2["A3"] = "Annualized revenue"
ws2["A3"].font = section_font
ws2.merge_cells("A3:E3")

ws2.cell(row=4, column=1, value="Scenario").font = hdr_font
ws2.cell(row=4, column=1).fill = hdr_fill
for col, n in enumerate(SCALES, start=2):
    c = ws2.cell(row=4, column=col, value=f"{n:,} creators")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

ws2.cell(row=5, column=1, value="Conservative").font = Font(bold=True)
for col, n in enumerate(SCALES, start=2):
    cell = ws2.cell(row=5, column=col, value=conservative_results[n].annualized_revenue)
    cell.number_format = money_format
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

ws2.cell(row=6, column=1, value="Base").font = Font(bold=True)
for col, n in enumerate(SCALES, start=2):
    cell = ws2.cell(row=6, column=col, value=base_results[n].annualized_revenue)
    cell.number_format = money_format
    cell.alignment = Alignment(horizontal="right")
    cell.fill = PatternFill("solid", fgColor=LIGHT_NEON)

ws2.cell(row=7, column=1, value="Conservative as % of Base").font = Font(italic=True)
for col, n in enumerate(SCALES, start=2):
    pct = conservative_results[n].annualized_revenue / base_results[n].annualized_revenue
    cell = ws2.cell(row=7, column=col, value=pct)
    cell.number_format = "0.0%"
    cell.alignment = Alignment(horizontal="right")

# Operating profit
ws2["A10"] = "Annualized operating profit"
ws2["A10"].font = section_font
ws2.merge_cells("A10:E10")

ws2.cell(row=11, column=1, value="Scenario").font = hdr_font
ws2.cell(row=11, column=1).fill = hdr_fill
for col, n in enumerate(SCALES, start=2):
    c = ws2.cell(row=11, column=col, value=f"{n:,} creators")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

ws2.cell(row=12, column=1, value="Conservative").font = Font(bold=True)
for col, n in enumerate(SCALES, start=2):
    cell = ws2.cell(row=12, column=col, value=conservative_results[n].annualized_profit)
    cell.number_format = money_format
    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)

ws2.cell(row=13, column=1, value="Base").font = Font(bold=True)
for col, n in enumerate(SCALES, start=2):
    cell = ws2.cell(row=13, column=col, value=base_results[n].annualized_profit)
    cell.number_format = money_format
    cell.fill = PatternFill("solid", fgColor=LIGHT_NEON)

# Per-creator metrics
ws2["A16"] = "Operating profit per creator per year"
ws2["A16"].font = section_font
ws2.merge_cells("A16:E16")

ws2.cell(row=17, column=1, value="Scenario").font = hdr_font
ws2.cell(row=17, column=1).fill = hdr_fill
for col, n in enumerate(SCALES, start=2):
    c = ws2.cell(row=17, column=col, value=f"{n:,} creators")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

ws2.cell(row=18, column=1, value="Conservative").font = Font(bold=True)
for col, n in enumerate(SCALES, start=2):
    cell = ws2.cell(row=18, column=col, value=conservative_results[n].annualized_profit / n)
    cell.number_format = money_format

ws2.cell(row=19, column=1, value="Base").font = Font(bold=True)
for col, n in enumerate(SCALES, start=2):
    cell = ws2.cell(row=19, column=col, value=base_results[n].annualized_profit / n)
    cell.number_format = money_format

ws2.column_dimensions["A"].width = 32
for col in "BCDE":
    ws2.column_dimensions[col].width = 18

# ─── Sheet 3: What differs ───
ws3 = wb.create_sheet("Assumption deltas")

ws3["A1"] = "What's different between Base and Conservative"
ws3["A1"].font = Font(name="Inter", size=18, bold=True, color=PINK)
ws3.merge_cells("A1:D1")

# Header
for col, label in enumerate(["Assumption", "Base", "Conservative", "Change"], start=1):
    c = ws3.cell(row=3, column=col, value=label)
    c.font = hdr_font
    c.fill = hdr_fill

deltas = [
    ("ACTIVATION RATES (% of registered who actually order)",),
    ("Maker activation", "30%", "15%", "Halved"),
    ("Builder activation", "70%", "35%", "Halved"),
    ("Agency activation", "90%", "45%", "Halved"),
    ("ORDERS / ACTIVE CREATOR / MONTH",),
    ("Maker orders/mo", "5", "2", "Halved"),
    ("Builder orders/mo", "80", "40", "Halved"),
    ("Agency orders/mo", "400", "200", "Halved"),
    ("BULK RUNS / ACTIVE CREATOR / MONTH",),
    ("Maker bulk runs/mo", "0.1", "0.05", "Halved"),
    ("Builder bulk runs/mo", "0.5", "0.25", "Halved"),
    ("Agency bulk runs/mo", "2.0", "1.0", "Halved"),
    ("VELOCITY DISTRIBUTION (% in Tier 1 — stuck at floor)",),
    ("Maker T1 share", "90%", "95%", "+5pp stuck"),
    ("Builder T1 share", "55%", "70%", "+15pp stuck"),
    ("Agency T1 share", "30%", "50%", "+20pp stuck"),
    ("WHAT STAYS IDENTICAL",),
    ("Pricing (subscription + velocity fees)", "Same", "Same", "Locked"),
    ("Partner tier mix / fees", "Same", "Same", "Locked"),
    ("Order economics ($30 retail / $8 wholesale)", "Same", "Same", "Locked"),
    ("Bulk run sizes", "Same", "Same", "Locked"),
    ("Variable costs (Stripe / AI / Shutterstock)", "Same", "Same", "Locked"),
    ("Fixed opex per scale", "Same", "Same", "Locked"),
]

row_i = 4
for item in deltas:
    if len(item) == 1:
        ws3.cell(row=row_i, column=1, value=item[0])
        ws3.cell(row=row_i, column=1).font = section_font
        ws3.cell(row=row_i, column=1).fill = section_fill
        ws3.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
    else:
        for col, val in enumerate(item, start=1):
            ws3.cell(row=row_i, column=col, value=val)
    row_i += 1

ws3.column_dimensions["A"].width = 45
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 14
ws3.column_dimensions["D"].width = 18

# ─── Sheet 4: Honest take ───
ws4 = wb.create_sheet("Honest take")
ws4["A1"] = "Honest take — what these two scenarios mean for planning"
ws4["A1"].font = Font(name="Inter", size=18, bold=True, color=PINK)
ws4.column_dimensions["A"].width = 110

notes = [
    "",
    "HOW TO READ EACH SCENARIO",
    "Conservative: this is what year one is likely to actually look like. Realistic activation, realistic order volume, most",
    "  creators never graduate past Velocity Tier 1. Use this for cash-flow planning, hiring runway, board updates.",
    "",
    "Base: this is the 'we found PMF and the velocity-tier mechanic is working' scenario. Realistic at month 18-30 IF beta",
    "  cohort 1 converts well, IF partner network scales, IF marketplace activation lands. Use this for fundraising upside",
    "  narrative (with explicit disclaimers) and internal stretch targets.",
    "",
    "WHAT THE GAP TELLS YOU",
    "Conservative is roughly 33-35% of Base ARR at every scale.",
    "→ The pricing model (velocity tiers) works the same in both — the GAP is entirely activation + order velocity + tier progression.",
    "→ Said differently: your platform's TAM is the same; the question is how fast creators move up the funnel.",
    "→ That's an EXECUTION lever (onboarding quality + creator success function + tier-upgrade prompts), not a PRICING lever.",
    "",
    "REASONABLE TARGETS",
    "Year 1 (100-1,000 creators): expect Conservative numbers. $95k-$1M ARR.",
    "Year 2 (1k-10k creators): expect mid-point between Conservative and Base. $1M-$10M ARR.",
    "Year 3 (10k-30k creators): expect Base numbers if execution holds. $25M-$85M ARR.",
    "",
    "USE THESE TO STRESS-TEST FUNDRAISE",
    "At 1,000 creators, Conservative says $1M ARR, Base says $2.9M. If you're raising on a $3M revenue projection, you need",
    "to be ready to defend the gap. Investors will model it themselves — better to show the Conservative case proactively",
    "and explain what closes the gap (Builder activation rising 35% → 70%, velocity-tier graduation accelerating).",
    "",
    "BURN RUNWAY MATH",
    "Conservative at 100 creators: ~$95k ARR. If you're burning $300k/yr at this stage, runway is the question.",
    "Base at 100 creators: $272k ARR. Approaching profitability.",
    "→ At pre-PMF, the difference between these scenarios decides whether you're raising defensively or from strength.",
    "",
    "DON'T DO THIS",
    "- Don't average the two scenarios for planning. They represent different worlds.",
    "- Don't pitch Conservative externally as 'realistic' — that signals you don't believe in the model.",
    "- Don't pitch Base externally as 'projection' — that's overpromising on assumptions you haven't validated.",
    "- Do: use Conservative for runway, Base for upside, and report the WIDER number on each axis (highest unit cost, lowest",
    "  activation) when stress-testing.",
]
for i, line in enumerate(notes, start=2):
    cell = ws4.cell(row=i, column=1, value=line)
    if line.startswith("→") or line.endswith("→"):
        cell.font = Font(bold=True, color=PINK)
    elif line.isupper() and len(line) > 5:
        cell.font = Font(bold=True, color=INK_900)

wb.save("./iLaunchify_Financial_Simulation_Conservative_vs_Base.xlsx")
print("\nWrote: iLaunchify_Financial_Simulation_Conservative_vs_Base.xlsx")
