"""
Build the Rebuilt Tier Table as a clean Excel spreadsheet.
5 sheets: Tier Comparison · Velocity Overlay · Upgrade Triggers · Value Math · Changes vs Locked
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Brand colors (per design system memory)
PINK = "FF2E63"
INK_900 = "111111"
INK_700 = "3A3A3A"
INK_500 = "888888"
CREAM = "F3EFE8"
NEON = "B5FF3D"
LIGHT_PINK = "FFE0EB"
LIGHT_NEON = "EFFFCC"
LIGHT_GREY = "F8F8F8"
WHITE = "FFFFFF"

# Per-tier color coding
MAKER_FILL = LIGHT_GREY
BUILDER_FILL = LIGHT_PINK
AGENCY_FILL = LIGHT_NEON

# Style helpers
title_font = Font(name="Inter", size=20, bold=True, color=PINK)
subtitle_font = Font(name="Inter", size=11, italic=True, color=INK_500)
section_font = Font(name="Inter", size=13, bold=True, color=INK_900)
hdr_font = Font(name="Inter", size=12, bold=True, color=WHITE)
tier_hdr_font = Font(name="Inter", size=14, bold=True, color=INK_900)
body_font = Font(name="Inter", size=11, color=INK_700)
emphasis_font = Font(name="Inter", size=11, bold=True, color=INK_900)
small_italic = Font(name="Inter", size=10, italic=True, color=INK_500)

hdr_fill = PatternFill("solid", fgColor=INK_900)
section_fill = PatternFill("solid", fgColor=CREAM)
yes_fill = PatternFill("solid", fgColor=LIGHT_NEON)
no_fill = PatternFill("solid", fgColor=LIGHT_GREY)

thin = Side(border_style="thin", color="E5E5E5")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_tier_header(ws, row, label, fill_color):
    """Style the 3-tier column header row."""
    cells = [
        (1, "Feature / Capability", INK_900, WHITE),
        (2, "Maker", INK_900, MAKER_FILL),
        (3, "Builder", INK_900, BUILDER_FILL),
        (4, "Agency", INK_900, AGENCY_FILL),
    ]
    for col, text, font_color, fill in cells:
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Inter", size=13, bold=True, color=font_color)
        c.fill = PatternFill("solid", fgColor=fill)
        c.alignment = center
        c.border = border


def feature_row(ws, row, label, maker, builder, agency, *, emphasize=False, note=None):
    """Write one feature row, color-coded for ✓ / —."""
    ws.cell(row=row, column=1, value=label).font = (emphasis_font if emphasize else body_font)
    ws.cell(row=row, column=1).alignment = left
    ws.cell(row=row, column=1).border = border
    if note:
        ws.cell(row=row, column=1).comment = None  # not adding comments to avoid clutter

    for col, val in enumerate([maker, builder, agency], start=2):
        c = ws.cell(row=row, column=col, value=val)
        c.alignment = center
        c.border = border
        c.font = body_font
        if val == "✓":
            c.fill = yes_fill
            c.font = Font(name="Inter", size=12, bold=True, color=INK_900)
        elif val == "—":
            c.fill = no_fill
            c.font = Font(name="Inter", size=11, color=INK_500)


wb = Workbook()

# ═══════════════════════════════════════════════════════════════════
# SHEET 1 — Rebuilt tier comparison
# ═══════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Tier Comparison"

ws["A1"] = "iLaunchify — Rebuilt Tier Model (V1.5 corrected)"
ws["A1"].font = title_font
ws.merge_cells("A1:D1")

ws["A2"] = ("Subscription gates features iLaunchify pays fixed cost to deliver. "
            "Never gates the sales mechanisms (bulk or on-demand). "
            "Velocity tier sits on top of subscription tier.")
ws["A2"].font = subtitle_font
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.merge_cells("A2:D2")
ws.row_dimensions[2].height = 36

# Tier headers
style_tier_header(ws, 4, None, None)

# Row data — every feature
row_i = 5

# Section: Audience
sec = ws.cell(row=row_i, column=1, value="WHO IT'S FOR")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i,
    "Target creator",
    "Evaluating their first bulk run",
    "Real CPG business operator",
    "Multi-brand operator / agency",
    emphasize=True)
row_i += 1

feature_row(ws, row_i,
    "Goal for iLaunchify",
    "Get them to first bulk order fast",
    "Retain + grow per-creator volume",
    "Portfolio scale + sales-touched")
row_i += 1

# Section: Pricing
row_i += 1
sec = ws.cell(row=row_i, column=1, value="PRICING")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i, "Monthly subscription", "$0", "$79/mo", "$249/mo", emphasize=True)
row_i += 1
feature_row(ws, row_i, "Velocity tier BASE fee", "15%", "10%", "7%")
row_i += 1
feature_row(ws, row_i, "Velocity tier FLOOR (T5)", "7%", "5%", "5% (was 3% — corrected)")
row_i += 1

# Section: Sales mechanisms
row_i += 1
sec = ws.cell(row=row_i, column=1, value="SALES MECHANISMS (universal — never gated)")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i, "Bulk orders", "✓", "✓", "✓", emphasize=True)
row_i += 1
feature_row(ws, row_i, "On-demand drop-ship", "✓", "✓", "✓", emphasize=True)
row_i += 1
feature_row(ws, row_i, "Unlimited products", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Velocity tier dynamic discount", "✓", "✓", "✓")
row_i += 1

# Section: Design Studio
row_i += 1
sec = ws.cell(row=row_i, column=1, value="DESIGN STUDIO + COMPLIANCE")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i, "Full canvas + dieline editing", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Compliance scan + label rendering", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Nutrition / Supplement / Drug / AAFCO panels", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "FDA claim auto-suggestion", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Custom fonts pinned to brand", "—", "✓", "✓")
row_i += 1

# Section: Asset library
row_i += 1
sec = ws.cell(row=row_i, column=1, value="ASSET LIBRARY (4-layer)")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i, "Layer 1 — Compliance graphics (free)", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Layer 2 — Unsplash + Pexels (free)", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Layer 2 — Shutterstock premium (paid)", "—", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Layer 3 — Iconify vectors (free)", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Layer 3 — Curated supplement vectors", "✓", "✓", "✓")
row_i += 1

# Section: AI
row_i += 1
sec = ws.cell(row=row_i, column=1, value="AI FEATURES (cost-bearing)")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i, "AI Recipe Parser (Anthropic)", "—", "50 / month", "200 / month")
row_i += 1
feature_row(ws, row_i, "AI image generation (Replicate)", "—", "20 / month", "100 / month")
row_i += 1
feature_row(ws, row_i, "Premium AI overage (per gen)", "n/a", "$1.50", "$1.50")
row_i += 1

# Section: Channels
row_i += 1
sec = ws.cell(row=row_i, column=1, value="CHANNEL INTEGRATIONS (per-integration cost)")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i, "Shopify push", "—", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "TikTok Shop push", "—", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Amazon push", "—", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "ClickFunnels integration", "—", "✓", "✓")
row_i += 1

# Section: Brand management
row_i += 1
sec = ws.cell(row=row_i, column=1, value="BRAND + PORTFOLIO MANAGEMENT")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i, "Brands per account", "1", "3", "Unlimited", emphasize=True)
row_i += 1
feature_row(ws, row_i, "Multi-brand dashboard", "—", "—", "✓")
row_i += 1
feature_row(ws, row_i, "Custom domain", "—", "—", "✓")
row_i += 1
feature_row(ws, row_i, "Team seats", "1", "1", "V1.5+ Creator Team")
row_i += 1

# Section: Account services
row_i += 1
sec = ws.cell(row=row_i, column=1, value="ACCOUNT SERVICES")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i, "Dedicated CSM", "—", "—", "✓ at $5k/mo+ usage")
row_i += 1
feature_row(ws, row_i, "Custom Stripe contract pricing", "—", "—", "✓ sales-touched")
row_i += 1
feature_row(ws, row_i, "Priority support", "Standard", "Standard", "Priority queue")
row_i += 1

# Section: Order add-ons
row_i += 1
sec = ws.cell(row=row_i, column=1, value="ORDER ADD-ONS")
sec.font = section_font
sec.fill = section_fill
ws.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=4)
row_i += 1

feature_row(ws, row_i, "Bulk escrow opt-in (0.5% × order)", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Partner-listed accessories (when offered)", "✓", "✓", "✓")
row_i += 1
feature_row(ws, row_i, "Float income on held balances", "auto-applied", "auto-applied", "auto-applied")
row_i += 1

# Column widths
ws.column_dimensions["A"].width = 50
for col in "BCD":
    ws.column_dimensions[col].width = 26

# Freeze first column + header row
ws.freeze_panes = "B5"

# ═══════════════════════════════════════════════════════════════════
# SHEET 2 — Velocity Tier Overlay
# ═══════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Velocity Overlay")

ws2["A1"] = "Velocity tier overlay (V1.5 corrected with 5% floor)"
ws2["A1"].font = title_font
ws2.merge_cells("A1:F1")

ws2["A2"] = ("Per-SKU 30-day rolling volume. Bulk volume counts toward on-demand tier (cross-pollination). "
             "Lower-of pricing. Samples always at Tier 1 and don't accrue.")
ws2["A2"].font = subtitle_font
ws2["A2"].alignment = Alignment(wrap_text=True)
ws2.merge_cells("A2:F2")
ws2.row_dimensions[2].height = 36

# Velocity table
hdr_row = 4
headers = ["Velocity Tier", "30-day SKU volume", "Maker fee %", "Builder fee %", "Agency fee %", "Notes"]
for col, h in enumerate(headers, start=1):
    c = ws2.cell(row=hdr_row, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = center
    c.border = border

velocity_data = [
    ("Tier 1", "0 – 50 units",       15.0, 10.0, 7.0, "Default floor for new SKUs and samples"),
    ("Tier 2", "51 – 200 units",     13.0, 8.5,  6.0, ""),
    ("Tier 3", "201 – 500 units",    11.0, 7.0,  5.0, ""),
    ("Tier 4", "501 – 1,000 units",   9.0, 6.0,  5.0, "Agency floor binds (was 4%)"),
    ("Tier 5", "1,000+ units",        7.0, 5.0,  5.0, "Agency floor binds (was 3%)"),
]
for i, (tier, vol, m, b, a, note) in enumerate(velocity_data, start=hdr_row + 1):
    ws2.cell(row=i, column=1, value=tier).font = emphasis_font
    ws2.cell(row=i, column=2, value=vol).font = body_font
    ws2.cell(row=i, column=3, value=m/100).number_format = "0.0%"
    ws2.cell(row=i, column=4, value=b/100).number_format = "0.0%"
    ws2.cell(row=i, column=5, value=a/100).number_format = "0.0%"
    ws2.cell(row=i, column=6, value=note).font = small_italic
    for col in range(1, 7):
        ws2.cell(row=i, column=col).alignment = center if col != 6 else left
        ws2.cell(row=i, column=col).border = border
    # Color-code the rows
    if i == hdr_row + 1:
        for col in range(3, 6):
            ws2.cell(row=i, column=col).fill = PatternFill("solid", fgColor=LIGHT_PINK)
    if i == hdr_row + 5:  # Tier 5
        for col in range(3, 6):
            ws2.cell(row=i, column=col).fill = PatternFill("solid", fgColor=LIGHT_NEON)

# Effective fees row
eff_row = hdr_row + 7
ws2.cell(row=eff_row, column=1, value="Effective fee (volume-weighted)").font = emphasis_font
ws2.cell(row=eff_row, column=1).fill = section_fill
ws2.cell(row=eff_row, column=2, value="across V1.5 distribution").font = small_italic
ws2.cell(row=eff_row, column=2).fill = section_fill
ws2.cell(row=eff_row, column=3, value=0.1476).number_format = "0.00%"
ws2.cell(row=eff_row, column=4, value=0.0892).number_format = "0.00%"
ws2.cell(row=eff_row, column=5, value=0.0583).number_format = "0.00%"
for col in range(1, 6):
    ws2.cell(row=eff_row, column=col).fill = section_fill
    ws2.cell(row=eff_row, column=col).border = border
    ws2.cell(row=eff_row, column=col).alignment = center

# Rules section
rules_row = eff_row + 3
ws2.cell(row=rules_row, column=1, value="LOCKED RULES").font = section_font
ws2.cell(row=rules_row, column=1).fill = section_fill
ws2.merge_cells(start_row=rules_row, start_column=1, end_row=rules_row, end_column=6)

rules = [
    "1. Per-SKU tracking — velocity is per (creator × ProductTemplate), not portfolio-aggregated",
    "2. Cross-pollination — bulk volume counts toward on-demand velocity tier",
    "3. Lower-of pricing — if bulk-quoted fee > current on-demand tier fee, lower fee applies",
    "4. Samples always at Tier 1 — no velocity discount, no accrual toward future tier",
    "5. Floor at 5% (corrected) — defends margin at Agency Tier 4 and Tier 5",
    "6. Admin-tuneable — VelocityTierThreshold values stored in DB, editable in /admin/tiers",
]
for i, rule in enumerate(rules, start=rules_row + 1):
    ws2.cell(row=i, column=1, value=rule).font = body_font
    ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

# Column widths
ws2.column_dimensions["A"].width = 22
ws2.column_dimensions["B"].width = 28
for col in "CDE":
    ws2.column_dimensions[col].width = 16
ws2.column_dimensions["F"].width = 32

# ═══════════════════════════════════════════════════════════════════
# SHEET 3 — Upgrade Triggers
# ═══════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Upgrade Triggers")

ws3["A1"] = "Upgrade triggers — soft nudges at moments of demonstrated need"
ws3["A1"].font = title_font
ws3.merge_cells("A1:C1")

ws3["A2"] = ("Triggers gate FEATURES creators want to unlock, not transactions they want to run. "
             "Language emphasizes UNLOCK, not COST. Bulk orders are never an upgrade trigger.")
ws3["A2"].font = subtitle_font
ws3["A2"].alignment = Alignment(wrap_text=True)
ws3.merge_cells("A2:C2")
ws3.row_dimensions[2].height = 36

# Maker → Builder
row_i = 4
ws3.cell(row=row_i, column=1, value="MAKER → BUILDER ($79/mo) triggers").font = section_font
ws3.cell(row=row_i, column=1).fill = section_fill
ws3.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=3)
row_i += 1

maker_to_builder = [
    ("Trigger event", "Prompt headline", "What unlocks"),
    ("2nd brand created", "Manage multiple brands under one account", "Up to 3 brands"),
    ("First channel push attempt", "Push your products to Shopify / TikTok Shop", "Channel integrations"),
    ("First AI Recipe Parser click", "Parse recipes from any source with AI", "50 parses / month"),
    ("First Shutterstock photo search", "Access 350M premium photos", "Shutterstock catalog"),
    ("First custom font search", "Pin custom fonts to your brand identity", "Custom font management"),
    ("4th product created", "Manage your growing portfolio with Builder", "Builder features"),
    ("First image gen attempt", "Generate custom images for your packaging", "20 AI images / month"),
]
for r in maker_to_builder:
    if r[0] == "Trigger event":
        for col, val in enumerate(r, start=1):
            c = ws3.cell(row=row_i, column=col, value=val)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = border
    else:
        for col, val in enumerate(r, start=1):
            c = ws3.cell(row=row_i, column=col, value=val)
            c.font = body_font
            c.alignment = left
            c.border = border
            if col == 1:
                c.font = emphasis_font
    row_i += 1

# Builder → Agency
row_i += 1
ws3.cell(row=row_i, column=1, value="BUILDER → AGENCY ($249/mo) triggers").font = section_font
ws3.cell(row=row_i, column=1).fill = section_fill
ws3.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=3)
row_i += 1

builder_to_agency = [
    ("Trigger event", "Prompt headline", "What unlocks"),
    ("4th brand created", "Multi-brand management is built for portfolios", "Unlimited brands + dashboard"),
    ("AI quota exhausted 2+ months", "Agency includes 4× the AI quota", "200 parses + 100 images / month"),
    ("30+ orders/day sustained", "Agency includes dedicated CSM at this volume", "CSM at $5k/mo+ usage"),
    ("Custom domain request", "Agency tier includes custom domain", "Custom domain support"),
    ("Team seat invitation", "Add teammates with Agency", "Creator Team model (V1.5+)"),
    ("Premium AI overage 3+ months", "Agency's higher cap is more cost-effective", "Higher base quota"),
]
for r in builder_to_agency:
    if r[0] == "Trigger event":
        for col, val in enumerate(r, start=1):
            c = ws3.cell(row=row_i, column=col, value=val)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = border
    else:
        for col, val in enumerate(r, start=1):
            c = ws3.cell(row=row_i, column=col, value=val)
            c.font = body_font
            c.alignment = left
            c.border = border
            if col == 1:
                c.font = emphasis_font
    row_i += 1

# Not a trigger (anti-list)
row_i += 1
ws3.cell(row=row_i, column=1, value="WHAT IS NOT AN UPGRADE TRIGGER (intentionally)").font = section_font
ws3.cell(row=row_i, column=1).fill = section_fill
ws3.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=3)
row_i += 1

anti = [
    ("First bulk order", "Bulk is the platform's core sales mechanism — never gated"),
    ("Higher-volume bulk run", "Velocity tier rewards volume regardless of subscription"),
    ("On-demand order", "Universal access to drop-ship fulfillment"),
    ("Marketplace browsing", "Free for all creators"),
    ("Compliance scan", "Required everywhere — iLaunchify covers the cost"),
    ("Sample order", "At-cost, always Tier 1 fee, never gated"),
]
for trigger, why in anti:
    ws3.cell(row=row_i, column=1, value=trigger).font = emphasis_font
    ws3.cell(row=row_i, column=2, value=why).font = body_font
    ws3.merge_cells(start_row=row_i, start_column=2, end_row=row_i, end_column=3)
    for col in (1, 2):
        ws3.cell(row=row_i, column=col).border = border
        ws3.cell(row=row_i, column=col).alignment = left
    row_i += 1

ws3.column_dimensions["A"].width = 36
ws3.column_dimensions["B"].width = 48
ws3.column_dimensions["C"].width = 36

# ═══════════════════════════════════════════════════════════════════
# SHEET 4 — Value math
# ═══════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Value Math")

ws4["A1"] = "Why each tier upgrade is rational for the creator"
ws4["A1"].font = title_font
ws4.merge_cells("A1:E1")

ws4["A2"] = ("Value/cost ratio per upgrade. Numbers are perceived monthly value at typical usage. "
             "Target: ≥2× value/cost for upgrade to feel obvious.")
ws4["A2"].font = subtitle_font
ws4["A2"].alignment = Alignment(wrap_text=True)
ws4.merge_cells("A2:E2")
ws4.row_dimensions[2].height = 36

# Maker → Builder
row_i = 4
ws4.cell(row=row_i, column=1, value="MAKER → BUILDER (+$79/mo)").font = section_font
ws4.cell(row=row_i, column=1).fill = section_fill
ws4.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=5)
row_i += 1

headers = ["Value component", "Description", "Monthly value", "Why this matters"]
for col, h in enumerate(headers, start=1):
    c = ws4.cell(row=row_i, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = center
    c.border = border
row_i += 1

mb_value = [
    ("AI Recipe Parser quota",     "50 parses × $0.30 cost (we eat)",         15,   "iLaunchify covers Anthropic cost"),
    ("AI image generation quota",  "20 gens × $0.05 cost (we eat)",            1,   "Replicate cost covered"),
    ("Shutterstock catalog",       "$300/mo enterprise seat amortized",       30,   "Premium stock library access"),
    ("Channel push (Shopify+TT)",  "Integrations we maintain",                50,   "Vs paying for Shopify dropship apps"),
    ("Custom fonts pinned",        "Brand identity sophistication",           20,   "Pro brand consistency"),
    ("Velocity base 15% → 10%",    "5pp fee compression × ~$240 fees/mo",    120,   "Direct fee savings, recurring"),
    ("Multi-brand (up to 3)",      "Manage portfolio under one account",      30,   "Avoids creating separate accounts"),
]
total_value_mb = 0
for component, desc, val, why in mb_value:
    ws4.cell(row=row_i, column=1, value=component).font = emphasis_font
    ws4.cell(row=row_i, column=2, value=desc).font = body_font
    ws4.cell(row=row_i, column=3, value=val).number_format = '"$"#,##0'
    ws4.cell(row=row_i, column=3).font = body_font
    ws4.cell(row=row_i, column=4, value=why).font = small_italic
    for col in range(1, 5):
        ws4.cell(row=row_i, column=col).border = border
        ws4.cell(row=row_i, column=col).alignment = left if col != 3 else Alignment(horizontal="right")
    total_value_mb += val
    row_i += 1

# Total + ratio
ws4.cell(row=row_i, column=1, value="TOTAL perceived monthly value").font = emphasis_font
ws4.cell(row=row_i, column=1).fill = section_fill
ws4.cell(row=row_i, column=3, value=total_value_mb).number_format = '"$"#,##0'
ws4.cell(row=row_i, column=3).font = Font(bold=True, size=12)
ws4.cell(row=row_i, column=3).fill = section_fill
row_i += 1
ws4.cell(row=row_i, column=1, value="Builder cost").font = emphasis_font
ws4.cell(row=row_i, column=3, value=79).number_format = '"$"#,##0'
ws4.cell(row=row_i, column=3).font = body_font
row_i += 1
ws4.cell(row=row_i, column=1, value="VALUE / COST RATIO").font = emphasis_font
ws4.cell(row=row_i, column=1).fill = PatternFill("solid", fgColor=LIGHT_NEON)
ws4.cell(row=row_i, column=3, value=total_value_mb / 79).number_format = '0.0"×"'
ws4.cell(row=row_i, column=3).font = Font(bold=True, size=14, color=PINK)
ws4.cell(row=row_i, column=3).fill = PatternFill("solid", fgColor=LIGHT_NEON)
row_i += 1

# Builder → Agency
row_i += 2
ws4.cell(row=row_i, column=1, value="BUILDER → AGENCY (+$170/mo incremental)").font = section_font
ws4.cell(row=row_i, column=1).fill = section_fill
ws4.merge_cells(start_row=row_i, start_column=1, end_row=row_i, end_column=5)
row_i += 1

for col, h in enumerate(headers, start=1):
    c = ws4.cell(row=row_i, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = center
    c.border = border
row_i += 1

ba_value = [
    ("Additional AI parser quota",  "+150 parses/mo × $0.30",                 45,   "Heavy users + agencies need volume"),
    ("Additional image gen quota",  "+80 gens/mo × $0.05",                     4,   "Larger creative volume"),
    ("Multi-brand dashboard",       "Portfolio operations UI",                 80,   "Real workflow for agencies"),
    ("Custom domain",               "ilaunchify-hosted brand domain",          25,   "Brand professionalism"),
    ("Velocity base 10% → 7%",      "3pp fee compression × ~$1500 fees/mo",  120,   "Direct fee savings at Agency volume"),
    ("Velocity floor 5% (T5)",      "Even at high velocity, 5% holds",         40,   "Margin defense for both sides"),
    ("Dedicated CSM (volume-gated)", "1:1 success at $5k/mo+ usage",           60,   "Worth ~$60/mo amortized"),
    ("Custom Stripe contract",      "Negotiated rates at scale",               30,   "Lower payment processing"),
]
total_value_ba = 0
for component, desc, val, why in ba_value:
    ws4.cell(row=row_i, column=1, value=component).font = emphasis_font
    ws4.cell(row=row_i, column=2, value=desc).font = body_font
    ws4.cell(row=row_i, column=3, value=val).number_format = '"$"#,##0'
    ws4.cell(row=row_i, column=3).font = body_font
    ws4.cell(row=row_i, column=4, value=why).font = small_italic
    for col in range(1, 5):
        ws4.cell(row=row_i, column=col).border = border
        ws4.cell(row=row_i, column=col).alignment = left if col != 3 else Alignment(horizontal="right")
    total_value_ba += val
    row_i += 1

ws4.cell(row=row_i, column=1, value="TOTAL incremental monthly value").font = emphasis_font
ws4.cell(row=row_i, column=1).fill = section_fill
ws4.cell(row=row_i, column=3, value=total_value_ba).number_format = '"$"#,##0'
ws4.cell(row=row_i, column=3).font = Font(bold=True, size=12)
ws4.cell(row=row_i, column=3).fill = section_fill
row_i += 1
ws4.cell(row=row_i, column=1, value="Builder → Agency incremental cost").font = emphasis_font
ws4.cell(row=row_i, column=3, value=170).number_format = '"$"#,##0'
ws4.cell(row=row_i, column=3).font = body_font
row_i += 1
ws4.cell(row=row_i, column=1, value="VALUE / COST RATIO").font = emphasis_font
ws4.cell(row=row_i, column=1).fill = PatternFill("solid", fgColor=LIGHT_NEON)
ws4.cell(row=row_i, column=3, value=total_value_ba / 170).number_format = '0.0"×"'
ws4.cell(row=row_i, column=3).font = Font(bold=True, size=14, color=PINK)
ws4.cell(row=row_i, column=3).fill = PatternFill("solid", fgColor=LIGHT_NEON)

ws4.column_dimensions["A"].width = 36
ws4.column_dimensions["B"].width = 38
ws4.column_dimensions["C"].width = 16
ws4.column_dimensions["D"].width = 40

# ═══════════════════════════════════════════════════════════════════
# SHEET 5 — Changes vs originally-locked spec
# ═══════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Changes vs Locked")

ws5["A1"] = "What changed in the rebuild — vs originally locked V1.5"
ws5["A1"].font = title_font
ws5.merge_cells("A1:D1")

ws5["A2"] = ("Three numeric changes + one philosophical clarification. Everything else stays locked.")
ws5["A2"].font = subtitle_font
ws5["A2"].alignment = Alignment(wrap_text=True)
ws5.merge_cells("A2:D2")

headers = ["Change", "Originally locked", "Rebuilt", "Reasoning"]
for col, h in enumerate(headers, start=1):
    c = ws5.cell(row=4, column=col, value=h)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = center
    c.border = border

changes = [
    ("Builder subscription", "$49/mo", "$79/mo",
     "60% raise. Sub is only 17% of customer LTV at Base; raising has minimal churn risk. Velocity tier discount is doing the real upgrade work."),
    ("Agency subscription", "$199/mo", "$249/mo",
     "25% raise. Agency value is more than 4× Builder; pricing better reflects delivered value."),
    ("Agency velocity tier 4", "4%", "5% (floor)",
     "Defends margin at the high-volume tail. 4% is below Stripe processing as a percentage."),
    ("Agency velocity tier 5", "3%", "5% (floor)",
     "3% turns Agency Tier 5 creators into break-even customers. 5% is still 28% off Agency base of 7%."),
    ("Maker bulk access (clarification)", "Implied unrestricted", "EXPLICITLY confirmed",
     "Bulk-first platform — bulk orders are never gated by subscription tier. Locked as core principle."),
    ("Tier philosophy (clarification)", "Implicit", "EXPLICITLY locked",
     "Subscription gates features we pay fixed cost for. Never gates sales mechanisms (bulk + on-demand)."),
]
for i, (change, was, now, why) in enumerate(changes, start=5):
    ws5.cell(row=i, column=1, value=change).font = emphasis_font
    ws5.cell(row=i, column=2, value=was).font = body_font
    ws5.cell(row=i, column=3, value=now).font = body_font
    ws5.cell(row=i, column=4, value=why).font = small_italic
    for col in range(1, 5):
        ws5.cell(row=i, column=col).border = border
        ws5.cell(row=i, column=col).alignment = left
    ws5.row_dimensions[i].height = 50

# What stays the same
ws5.cell(row=12, column=1, value="WHAT STAYS LOCKED (no change from V1.5 spec)").font = section_font
ws5.cell(row=12, column=1).fill = section_fill
ws5.merge_cells(start_row=12, start_column=1, end_row=12, end_column=4)

unchanged = [
    "3-tier structure: Maker / Builder / Agency",
    "Velocity tier dynamics: 5 tiers per subscription tier, 30-day rolling window",
    "Per-SKU velocity tracking (not per-Brand)",
    "Cross-pollination — bulk volume counts toward on-demand tier",
    "Lower-of pricing at bulk-quote time",
    "Samples always at Tier 1, no velocity accrual",
    "Partner-side fees: Verified 5% / Trusted 3.5% / Premier 2% (flat, no velocity in V1.5)",
    "Marketplace fulfillment-mode visual treatment",
    "Asset library 4-layer architecture",
    "Multi-component packaging schema",
    "Decoration method as first-class concept",
    "Dieline normalization workflow",
    "Compliance template Recipal-model UX",
    "All 8 V1 tracks remain as scoped",
]
for i, item in enumerate(unchanged, start=13):
    ws5.cell(row=i, column=1, value="•  " + item).font = body_font
    ws5.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

ws5.column_dimensions["A"].width = 32
ws5.column_dimensions["B"].width = 22
ws5.column_dimensions["C"].width = 22
ws5.column_dimensions["D"].width = 60

wb.save("./iLaunchify_Rebuilt_Tier_Model.xlsx")
print("Wrote: iLaunchify_Rebuilt_Tier_Model.xlsx")
