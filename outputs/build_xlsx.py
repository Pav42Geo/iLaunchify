"""Build iLaunchify_Financial_Simulation.xlsx from the JSON output."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("./financial_simulation_results.json") as f:
    data = json.load(f)

results = data["results"]
fees = data["effective_fees"]

wb = Workbook()

# Styles
PINK = "FF2E63"
INK_900 = "111111"
NEON = "B5FF3D"
CREAM = "F3EFE8"

hdr_font = Font(name="Inter", size=12, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=INK_900)
section_font = Font(name="Inter", size=14, bold=True, color=INK_900)
section_fill = PatternFill("solid", fgColor=CREAM)
money_format = '"$"#,##0'
pct_format = '0.00"%"'
thin = Side(border_style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ─────────────────────────────────────────────────────────────────────
# Sheet 1 — Summary results matrix
# ─────────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = "Summary"

ws["A1"] = "iLaunchify financial simulation — V1.5 locked model"
ws["A1"].font = Font(name="Inter", size=18, bold=True, color=PINK)
ws.merge_cells("A1:F1")

ws["A2"] = "Generated from financial_simulation.py. Tune assumptions in the script and rerun. See README sheet."
ws["A2"].font = Font(italic=True, size=10, color="666666")
ws.merge_cells("A2:F2")

ws["A4"] = "Effective creator fee % (volume-weighted across velocity tiers)"
ws["A4"].font = section_font
ws.merge_cells("A4:F4")

ws.append([])
ws["A5"] = "Maker"
ws["B5"] = fees["maker"] / 100
ws["A6"] = "Builder"
ws["B6"] = fees["builder"] / 100
ws["A7"] = "Agency"
ws["B7"] = fees["agency"] / 100
ws["A8"] = "Partner-side avg (mix-weighted)"
ws["B8"] = fees["partner_avg"] / 100

for r in range(5, 9):
    ws.cell(row=r, column=2).number_format = "0.00%"
    ws.cell(row=r, column=1).font = Font(bold=True)

# Results matrix
ws["A10"] = "Scenario matrix"
ws["A10"].font = section_font
ws.merge_cells("A10:F10")

header = ["Metric", "100 creators", "1,000 creators", "10,000 creators", "30,000 creators"]
ws.append([])  # blank row 11
for col, h in enumerate(header, start=1):
    cell = ws.cell(row=12, column=col, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center")

def row(label: str, key: str, fmt: str = money_format):
    nonlocal_row = ws.max_row + 1
    ws.cell(row=nonlocal_row, column=1, value=label).font = Font(bold=True if "Total" in label or "Annualized" in label else False)
    for col_idx, r in enumerate(results, start=2):
        cell = ws.cell(row=nonlocal_row, column=col_idx, value=r[key])
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right")
    return nonlocal_row

# User counts
row("Partner count", "partner_count", "#,##0")
row("Active Makers", "active_makers", "#,##0")
row("Active Builders", "active_builders", "#,##0")
row("Active Agencies", "active_agencies", "#,##0")

# Spacer
ws.append([])

# Subscription
ws.cell(row=ws.max_row + 1, column=1, value="SUBSCRIPTION REVENUE (monthly)").font = section_font
row("Subscription revenue", "subscription_revenue_monthly")

# Spacer
ws.append([])
ws.cell(row=ws.max_row, column=1, value="ON-DEMAND (monthly)").font = section_font
row("On-demand units", "on_demand_units_monthly", "#,##0")
row("On-demand gross retail", "on_demand_gross_revenue_monthly")
row("Creator-side platform fee", "on_demand_platform_fee_creator_side")
row("Partner-side platform fee", "on_demand_platform_fee_partner_side")

# Spacer
ws.append([])
ws.cell(row=ws.max_row, column=1, value="BULK (monthly)").font = section_font
row("Bulk units", "bulk_units_monthly", "#,##0")
row("Bulk partner cost (gross)", "bulk_gross_revenue_monthly")
row("Creator-side platform fee", "bulk_platform_fee_creator_side")
row("Partner-side platform fee", "bulk_platform_fee_partner_side")

# Costs
ws.append([])
ws.cell(row=ws.max_row, column=1, value="COSTS (monthly)").font = section_font
row("Stripe Connect (variable)", "stripe_costs_monthly")
row("AI Recipe Parser (Anthropic)", "ai_parser_costs_monthly")
row("Asset library (Shutterstock)", "asset_library_costs_monthly")
row("Fixed opex (eng+infra+ops)", "fixed_opex_monthly")

# Totals
ws.append([])
ws.cell(row=ws.max_row, column=1, value="BOTTOM LINE").font = section_font
row("Total platform take (monthly)", "total_platform_take_monthly")
row("Net revenue (after variable costs)", "net_revenue_monthly")
row("Operating profit (monthly)", "operating_profit_monthly")
row("Annualized revenue", "annualized_revenue")
row("Annualized operating profit", "annualized_profit")

# Highlight last 2 rows
for r_idx in (ws.max_row - 1, ws.max_row):
    for col_idx in range(1, 6):
        ws.cell(row=r_idx, column=col_idx).fill = PatternFill("solid", fgColor=CREAM)
        ws.cell(row=r_idx, column=col_idx).font = Font(bold=True, size=12)

# Column widths
ws.column_dimensions["A"].width = 42
for col in "BCDEF":
    ws.column_dimensions[col].width = 18

# ─────────────────────────────────────────────────────────────────────
# Sheet 2 — Assumptions
# ─────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Assumptions")

ws2["A1"] = "Assumptions — tune these in financial_simulation.py and rerun"
ws2["A1"].font = Font(name="Inter", size=16, bold=True, color=PINK)
ws2.merge_cells("A1:C1")

ws2["A2"] = "All inputs that drive the Summary sheet. These are aggressive base-case values."
ws2["A2"].font = Font(italic=True, size=10, color="666666")
ws2.merge_cells("A2:C2")

def section(ws, label, row):
    ws.cell(row=row, column=1, value=label).font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    return row + 1

row_idx = 4
row_idx = section(ws2, "Creator subscription tier mix", row_idx)
for tier, share in [("Maker", 0.70), ("Builder ($49/mo)", 0.25), ("Agency ($199/mo)", 0.05)]:
    ws2.cell(row=row_idx, column=1, value=tier)
    ws2.cell(row=row_idx, column=2, value=share).number_format = "0.0%"
    row_idx += 1

row_idx += 1
row_idx = section(ws2, "Creator activation rate (% who order monthly)", row_idx)
for tier, rate in [("Maker", 0.30), ("Builder", 0.70), ("Agency", 0.90)]:
    ws2.cell(row=row_idx, column=1, value=tier)
    ws2.cell(row=row_idx, column=2, value=rate).number_format = "0.0%"
    row_idx += 1

row_idx += 1
row_idx = section(ws2, "On-demand orders per active creator per month", row_idx)
for tier, n in [("Maker", 5), ("Builder", 80), ("Agency", 400)]:
    ws2.cell(row=row_idx, column=1, value=tier)
    ws2.cell(row=row_idx, column=2, value=n).number_format = "#,##0"
    row_idx += 1

row_idx += 1
row_idx = section(ws2, "Bulk runs per active creator per month + avg run size", row_idx)
for tier, runs, size in [("Maker", 0.1, 200), ("Builder", 0.5, 800), ("Agency", 2.0, 3000)]:
    ws2.cell(row=row_idx, column=1, value=tier)
    ws2.cell(row=row_idx, column=2, value=runs).number_format = "0.0"
    ws2.cell(row=row_idx, column=3, value=size).number_format = "#,##0"
    row_idx += 1

row_idx += 1
row_idx = section(ws2, "Velocity tier fee % (Maker / Builder / Agency)", row_idx)
ws2.cell(row=row_idx, column=1, value="Tier 1 (0-50)")
ws2.cell(row=row_idx, column=2, value="15% / 10% / 7%")
row_idx += 1
ws2.cell(row=row_idx, column=1, value="Tier 2 (51-200)")
ws2.cell(row=row_idx, column=2, value="13% / 8.5% / 6%")
row_idx += 1
ws2.cell(row=row_idx, column=1, value="Tier 3 (201-500)")
ws2.cell(row=row_idx, column=2, value="11% / 7% / 5%")
row_idx += 1
ws2.cell(row=row_idx, column=1, value="Tier 4 (501-1000)")
ws2.cell(row=row_idx, column=2, value="9% / 6% / 4%")
row_idx += 1
ws2.cell(row=row_idx, column=1, value="Tier 5 (1000+)")
ws2.cell(row=row_idx, column=2, value="7% / 5% / 3%")
row_idx += 1

row_idx += 1
row_idx = section(ws2, "Velocity tier distribution (% of active creators per sub tier)", row_idx)
ws2.cell(row=row_idx, column=1, value="Maker  → T1:90% T2:8% T3:2%")
row_idx += 1
ws2.cell(row=row_idx, column=1, value="Builder → T1:55% T2:25% T3:12% T4:6% T5:2%")
row_idx += 1
ws2.cell(row=row_idx, column=1, value="Agency  → T1:30% T2:30% T3:20% T4:12% T5:8%")
row_idx += 1

row_idx += 1
row_idx = section(ws2, "Order economics", row_idx)
for label, val in [
    ("On-demand retail AOV", "$30.00"),
    ("On-demand partner wholesale", "$8.00"),
    ("On-demand shipping est", "$4.50"),
    ("Bulk partner wholesale per unit", "$4.50"),
]:
    ws2.cell(row=row_idx, column=1, value=label)
    ws2.cell(row=row_idx, column=2, value=val)
    row_idx += 1

row_idx += 1
row_idx = section(ws2, "Partner fee % (mix-weighted on partner wholesale)", row_idx)
for tier, pct in [("Verified (70%)", "5.0%"), ("Trusted (25%)", "3.5%"), ("Premier (5%)", "2.0%")]:
    ws2.cell(row=row_idx, column=1, value=tier)
    ws2.cell(row=row_idx, column=2, value=pct)
    row_idx += 1

row_idx += 1
row_idx = section(ws2, "Other parameters", row_idx)
for label, val in [
    ("Creators per partner", "40"),
    ("Stripe Connect %", "0.25% + $0.25 / tx"),
    ("AI parser cost / parse (Anthropic)", "$0.30"),
    ("AI parses / Builder / month", "50"),
    ("AI parses / Agency / month", "200"),
    ("Shutterstock monthly", "$300 flat"),
]:
    ws2.cell(row=row_idx, column=1, value=label)
    ws2.cell(row=row_idx, column=2, value=val)
    row_idx += 1

row_idx += 1
row_idx = section(ws2, "Fixed opex by scale (eng + infra + tools)", row_idx)
for n, opex in [(100, "$25k/yr"), (1000, "$75k/yr"), (10_000, "$250k/yr"), (30_000, "$600k/yr")]:
    ws2.cell(row=row_idx, column=1, value=f"~{n:,} users")
    ws2.cell(row=row_idx, column=2, value=opex)
    row_idx += 1

ws2.column_dimensions["A"].width = 50
ws2.column_dimensions["B"].width = 22
ws2.column_dimensions["C"].width = 18

# ─────────────────────────────────────────────────────────────────────
# Sheet 3 — README + caveats
# ─────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("README")
ws3["A1"] = "README — how to read this simulation"
ws3["A1"].font = Font(name="Inter", size=16, bold=True, color=PINK)
ws3.merge_cells("A1:A1")
ws3.column_dimensions["A"].width = 120

notes = [
    "",
    "WHAT THIS IS",
    "Aggressive base-case financial projection of iLaunchify at four creator scales: 100, 1,000, 10,000, 30,000 active creators.",
    "Built against the V1.5 LOCKED pricing model: subscription tiers + on-demand creator/partner fees + velocity-tier discount layer + bulk fees.",
    "Outputs cover BOTH user types: creators (subscription + per-order fees) and partners (per-order wholesale fees).",
    "",
    "HOW THE NUMBERS ARE BUILT",
    "1. Total creators are split across Maker (70%) / Builder (25%) / Agency (5%) per typical SaaS freemium curves.",
    "2. Each subscription tier has an activation rate — Maker 30% / Builder 70% / Agency 90%.",
    "3. Active creators place on-demand AND bulk orders at tier-specific monthly volumes.",
    "4. The effective creator-side fee % is volume-weighted across the 5 velocity tiers per subscription tier.",
    "5. Partner-side fees flow on the partner wholesale value at a 70/25/5 Verified/Trusted/Premier mix.",
    "6. Fixed opex grows with scale (eng team + infra + ops).",
    "",
    "IMPORTANT CAVEATS",
    "These are AGGRESSIVE base-case assumptions. Real-world outcomes will likely be 30-60% of these projections for the first 12 months.",
    "Specifically optimistic: 80 orders/mo per active Builder (real distribution will be long-tail; median Builder may do 5-20).",
    "Specifically optimistic: 2 bulk runs/mo per Agency × 3,000 units = 72k units/Agency/year. Realistic Agency creators may do 1-3 bulk runs total per year.",
    "Specifically optimistic: Velocity distribution assumes creators move up the curve. Many will stay at Tier 1 forever.",
    "",
    "WHAT THE NUMBERS CAN BE USED FOR",
    "- Order of magnitude sanity-check on pricing model viability.",
    "- Fundraising narrative (with explicit assumption caveats).",
    "- Internal target-setting (what would scale look like if these assumptions held).",
    "",
    "WHAT THE NUMBERS CANNOT BE USED FOR",
    "- External investor pitch claims without disclaimers — these are projections not actuals (per Pavel's P2 fabricated-traction strip).",
    "- Predictive monthly revenue forecasting until you have real cohort data.",
    "",
    "TUNING",
    "Open financial_simulation.py, change the ASSUMPTIONS block at the top, run `python3 financial_simulation.py`,",
    "then rerun build_xlsx.py to refresh this workbook. All math is explicit, no hidden multipliers.",
    "",
    "SENSITIVITY GUIDANCE",
    "Most sensitive levers (small changes have big impact):",
    "  1. Builder activation rate (drops to 50% → revenue drops ~30%)",
    "  2. Orders per active Builder (cuts to 30 → revenue drops ~40%)",
    "  3. Velocity tier distribution at Agency (if T4+T5 share drops to 5%, agency fees up 1pt → revenue rises ~10%)",
    "",
    "Least sensitive levers:",
    "  - Asset library + AI parser costs (rounding error at scale)",
    "  - Partner tier mix (small impact since partner fees are small % of wholesale)",
]
for i, line in enumerate(notes, start=2):
    cell = ws3.cell(row=i, column=1, value=line)
    if line and not line.startswith(" ") and not line.startswith("-") and not line.startswith("1") and not line.startswith("2") and not line.startswith("3") and not line.startswith("4") and not line.startswith("5") and not line.startswith("6"):
        if line.isupper() or (line and line[0].isupper() and ":" not in line and len(line.split()) <= 6):
            cell.font = Font(bold=True, color=INK_900)

wb.save("/sessions/brave-affectionate-mayer/mnt/iLaunchify/outputs/iLaunchify_Financial_Simulation.xlsx")
print("Wrote: iLaunchify_Financial_Simulation.xlsx")
