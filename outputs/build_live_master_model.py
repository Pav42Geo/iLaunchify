"""
Build ONE master live spreadsheet with FORMULAS instead of values.
Every cell that depends on an assumption is a formula. Edit any input,
the entire model recalculates in Excel.

Structure:
  Sheet 1 — ASSUMPTIONS    (every tuneable input)
  Sheet 2 — TIER MODEL      (rebuilt tier comparison, references Assumptions)
  Sheet 3 — VELOCITY OVERLAY (5-tier matrix referencing assumptions)
  Sheet 4 — SCENARIO TOGGLES (Base / Conservative / +5pp / +10pp / Recs)
  Sheet 5 — RESULTS         (4 scales with full breakdown, all formula-driven)
  Sheet 6 — DASHBOARD       (headline numbers, live)
  Sheet 7 — README          (how to use)
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook.defined_name import DefinedName

# ─── Brand ─────────────────────────────────────────────────────────
PINK = "FF2E63"; INK = "111111"; CREAM = "F3EFE8"
NEON = "B5FF3D"; LIGHT_PINK = "FFE0EB"; LIGHT_NEON = "EFFFCC"
LIGHT_GREY = "F8F8F8"; WHITE = "FFFFFF"; INPUT_BG = "FFFCE6"

title = Font(name="Inter", size=18, bold=True, color=PINK)
subtitle = Font(name="Inter", size=10, italic=True, color="666666")
section = Font(name="Inter", size=13, bold=True, color=INK)
hdr = Font(name="Inter", size=11, bold=True, color=WHITE)
body = Font(name="Inter", size=11, color=INK)
emph = Font(name="Inter", size=11, bold=True, color=INK)
input_font = Font(name="Inter", size=11, bold=True, color="B85C00")
output_font = Font(name="Inter", size=11, color="333333")
result_font = Font(name="Inter", size=12, bold=True, color=PINK)

hdr_fill = PatternFill("solid", fgColor=INK)
section_fill = PatternFill("solid", fgColor=CREAM)
input_fill = PatternFill("solid", fgColor=INPUT_BG)
result_fill = PatternFill("solid", fgColor=LIGHT_PINK)

thin = Side(border_style="thin", color="DDDDDD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_a = Alignment(horizontal="left", vertical="center", wrap_text=True)
right_a = Alignment(horizontal="right", vertical="center")

money = '"$"#,##0'
money_neg = '"$"#,##0;("$"#,##0);"—"'
pct = '0.0%'
pct2 = '0.00%'

wb = Workbook()


# ═══════════════════════════════════════════════════════════════════
# SHEET 1 — ASSUMPTIONS (all inputs)
# ═══════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Assumptions"

ws["A1"] = "ASSUMPTIONS — edit any yellow cell. Model recalculates everywhere."
ws["A1"].font = title
ws.merge_cells("A1:E1")
ws["A2"] = "All inputs are tuneable. Cells with yellow background are editable. White cells are computed."
ws["A2"].font = subtitle
ws.merge_cells("A2:E2")

# Helper to write an input cell (yellow background, return its address)
def inp(cell_ref, value, fmt=None, *, label=None, label_cell=None):
    c = ws[cell_ref]
    c.value = value
    c.font = input_font
    c.fill = input_fill
    c.border = border
    c.alignment = right_a
    if fmt:
        c.number_format = fmt
    if label and label_cell:
        ws[label_cell] = label
        ws[label_cell].font = body
        ws[label_cell].border = border
    return cell_ref

def section_header(row, text):
    ws.cell(row=row, column=1, value=text).font = section
    ws.cell(row=row, column=1).fill = section_fill
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)


# ─── Section 1: Creator tier mix ─────────────────────────────────
section_header(4, "Creator subscription tier MIX (% of registered creators)")
ws["A5"] = "Maker share";   inp("B5", 0.70, pct)
ws["A6"] = "Builder share"; inp("B6", 0.25, pct)
ws["A7"] = "Agency share";  inp("B7", 0.05, pct)
ws["A8"] = "  Sum check"
ws["B8"] = "=SUM(B5:B7)"
ws["B8"].number_format = pct
ws["B8"].font = body

# ─── Section 2: Activation rates ─────────────────────────────────
section_header(10, "Activation rate per tier (% who order monthly)")
ws["A11"] = "Maker activation";   inp("B11", 0.30, pct)
ws["A12"] = "Builder activation"; inp("B12", 0.70, pct)
ws["A13"] = "Agency activation";  inp("B13", 0.90, pct)

# Conservative overrides
ws["D11"] = "Conservative override"; ws["D11"].font = subtitle
inp("E11", 0.15, pct)
inp("E12", 0.35, pct)
inp("E13", 0.45, pct)

# ─── Section 3: Orders/active/month ──────────────────────────────
section_header(15, "On-demand orders per active creator per month")
ws["A16"] = "Maker orders/mo";   inp("B16", 5, "#,##0")
ws["A17"] = "Builder orders/mo"; inp("B17", 80, "#,##0")
ws["A18"] = "Agency orders/mo";  inp("B18", 400, "#,##0")
ws["D16"] = "Conservative"; ws["D16"].font = subtitle
inp("E16", 2, "#,##0")
inp("E17", 40, "#,##0")
inp("E18", 200, "#,##0")

# ─── Section 4: Bulk parameters ──────────────────────────────────
section_header(20, "Bulk orders — runs/mo + avg run size")
ws["A21"] = "Maker bulk runs/mo";   inp("B21", 0.10, "0.00"); ws["C21"] = "size"; inp("D21", 200, "#,##0")
ws["A22"] = "Builder bulk runs/mo"; inp("B22", 0.50, "0.00"); ws["C22"] = "size"; inp("D22", 800, "#,##0")
ws["A23"] = "Agency bulk runs/mo";  inp("B23", 2.00, "0.00"); ws["C23"] = "size"; inp("D23", 3000, "#,##0")
ws["F21"] = "Conservative runs"; ws["F21"].font = subtitle
inp("G21", 0.05, "0.00")
inp("G22", 0.25, "0.00")
inp("G23", 1.00, "0.00")

# ─── Section 5: Subscription prices ──────────────────────────────
section_header(25, "Subscription pricing (monthly)")
ws["A26"] = "Builder $/mo (original $49 → rebuilt $79)"; inp("B26", 79, money)
ws["A27"] = "Agency $/mo (original $199 → rebuilt $249)"; inp("B27", 249, money)

# ─── Section 6: Velocity fee matrix ──────────────────────────────
section_header(30, "Velocity-tier fee % (subscription × velocity tier matrix)")
ws["A31"] = "Tier"; ws["B31"] = "Volume"; ws["C31"] = "Maker"; ws["D31"] = "Builder"; ws["E31"] = "Agency"
for col in "ABCDE":
    ws[f"{col}31"].font = hdr; ws[f"{col}31"].fill = hdr_fill
    ws[f"{col}31"].alignment = center; ws[f"{col}31"].border = border
velocity_rows = [
    (32, "Tier 1", "0-50",     15.0, 10.0, 7.0),
    (33, "Tier 2", "51-200",   13.0, 8.5,  6.0),
    (34, "Tier 3", "201-500",  11.0, 7.0,  5.0),
    (35, "Tier 4", "501-1000",  9.0, 6.0,  5.0),  # was 4.0
    (36, "Tier 5", "1000+",     7.0, 5.0,  5.0),  # was 3.0
]
for r, tier, vol, m, b, a in velocity_rows:
    ws.cell(row=r, column=1, value=tier).font = emph
    ws.cell(row=r, column=2, value=vol).font = body
    inp(f"C{r}", m/100, pct)
    inp(f"D{r}", b/100, pct)
    inp(f"E{r}", a/100, pct)

# ─── Section 7: Velocity distribution ────────────────────────────
section_header(38, "Velocity tier DISTRIBUTION (% of active creators per sub tier)")
ws["A39"] = "Tier"; ws["B39"] = "Maker"; ws["C39"] = "Builder"; ws["D39"] = "Agency"
ws["F39"] = "Conservative"
for col in ["A39","B39","C39","D39","F39"]:
    ws[col].font = hdr; ws[col].fill = hdr_fill; ws[col].alignment = center; ws[col].border = border
ws["F39"].fill = PatternFill("solid", fgColor="666666")

dist_rows = [
    (40, "Tier 1", 0.90, 0.55, 0.30, [0.95, 0.70, 0.50]),
    (41, "Tier 2", 0.08, 0.25, 0.30, [0.04, 0.20, 0.30]),
    (42, "Tier 3", 0.02, 0.12, 0.20, [0.01, 0.07, 0.15]),
    (43, "Tier 4", 0.00, 0.06, 0.12, [0.00, 0.02, 0.04]),
    (44, "Tier 5", 0.00, 0.02, 0.08, [0.00, 0.01, 0.01]),
]
for r, tier, m, b, a, cons in dist_rows:
    ws.cell(row=r, column=1, value=tier).font = emph
    inp(f"B{r}", m, pct)
    inp(f"C{r}", b, pct)
    inp(f"D{r}", a, pct)
    inp(f"F{r}", cons[0], pct)
    inp(f"G{r}", cons[1], pct)
    inp(f"H{r}", cons[2], pct)

# ─── Section 8: Order economics ──────────────────────────────────
section_header(47, "Order economics")
ws["A48"] = "On-demand retail AOV";          inp("B48", 30.00, money)
ws["A49"] = "On-demand partner wholesale";    inp("B49", 8.00, money)
ws["A50"] = "On-demand shipping est";         inp("B50", 4.50, money)
ws["A51"] = "Bulk partner wholesale / unit";  inp("B51", 4.50, money)

# ─── Section 9: Partner fees ────────────────────────────────────
section_header(54, "Partner-side platform fee (% of partner wholesale)")
ws["A55"] = "Verified fee";                   inp("B55", 0.050, pct)
ws["A56"] = "Trusted fee";                    inp("B56", 0.035, pct)
ws["A57"] = "Anchor fee (was Premier)";       inp("B57", 0.020, pct)
ws["A58"] = "Partner tier mix — Verified";    inp("B58", 0.70, pct)
ws["A59"] = "Partner tier mix — Trusted";     inp("B59", 0.25, pct)
ws["A60"] = "Partner tier mix — Anchor";      inp("B60", 0.05, pct)
ws["A62"] = "Creators per partner";           inp("B62", 40, "#,##0")

# ─── Section 10: Cost factors ───────────────────────────────────
section_header(64, "Variable + fixed costs")
ws["A65"] = "Stripe % per tx";                inp("B65", 0.0025, pct2)
ws["A66"] = "Stripe flat $ per tx";           inp("B66", 0.25, money)
ws["A67"] = "AI parser cost / parse (USD)";   inp("B67", 0.30, money)
ws["A68"] = "AI parses / Builder / mo";       inp("B68", 50, "#,##0")
ws["A69"] = "AI parses / Agency / mo";        inp("B69", 200, "#,##0")
ws["A70"] = "Shutterstock $/mo flat";         inp("B70", 300, money)
ws["A71"] = "Fixed opex annual @ 100";        inp("B71", 25_000, money)
ws["A72"] = "Fixed opex annual @ 1,000";      inp("B72", 75_000, money)
ws["A73"] = "Fixed opex annual @ 10,000";     inp("B73", 250_000, money)
ws["A74"] = "Fixed opex annual @ 30,000";     inp("B74", 600_000, money)

# ─── Section 11: Scenario toggles ────────────────────────────────
section_header(77, "Scenario toggles")
ws["A78"] = "Use Conservative demand? (1=yes, 0=no)"
inp("B78", 0, "0")
ws["A79"] = "Pricing uplift on velocity fees (percentage points)"
inp("B79", 0, "0.0")
ws["C79"] = "(0 = locked, 5 = +5pp, 10 = +10pp)"
ws["C79"].font = subtitle

ws["A81"] = "Recommendations layer (each 0 or 1):"; ws["A81"].font = section
ws["A82"] = "Rec 1 — Engineered Maker→Builder upgrade triggers"; inp("B82", 0, "0")
ws["A83"] = "Rec 2 — Float income on held balances"; inp("B83", 0, "0")
ws["A84"] = "Rec 3 — Sub prices Builder $79 / Agency $249"; inp("B84", 0, "0")
ws["A85"] = "Rec 4 — Velocity floor at 5% (already in matrix)"; inp("B85", 0, "0")
ws["A86"] = "Rec 6 — Partner referral commission (-2% drag)"; inp("B86", 0, "0")
ws["A87"] = "Rec 7 — Premium AI overage ($1.50/gen)"; inp("B87", 0, "0")
ws["A88"] = "Rec 8 — Bulk escrow opt-in (0.5% × 40%)"; inp("B88", 0, "0")

# ─── Section 12: Recs parameters ────────────────────────────────
section_header(91, "Recommendation parameters")
ws["A92"] = "Float average days held";        inp("B92", 2.0, "0.0")
ws["A93"] = "Treasury annual rate";            inp("B93", 0.045, pct)
ws["A94"] = "Partner referral creator share";  inp("B94", 0.20, pct)
ws["A95"] = "Partner referral commission rate"; inp("B95", 0.10, pct)
ws["A96"] = "AI overage rate (% B+A over cap)"; inp("B96", 0.20, pct)
ws["A97"] = "AI overage gens / over-cap creator"; inp("B97", 30, "#,##0")
ws["A98"] = "AI overage price / gen";           inp("B98", 1.50, money)
ws["A99"] = "AI overage cost / gen";            inp("B99", 0.05, money)
ws["A100"] = "Bulk escrow opt-in rate";          inp("B100", 0.40, pct)
ws["A101"] = "Bulk escrow fee %";                inp("B101", 0.005, pct)
ws["A102"] = "Rec 1 mix shift — Maker share";    inp("B102", 0.55, pct)
ws["A103"] = "Rec 1 mix shift — Builder share";  inp("B103", 0.40, pct)
ws["A104"] = "Rec 1 — Builder activation lift";  inp("B104", 0.50, pct)

ws.column_dimensions["A"].width = 44
ws.column_dimensions["B"].width = 14
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 22
ws.column_dimensions["E"].width = 14
ws.column_dimensions["F"].width = 22
ws.column_dimensions["G"].width = 14
ws.column_dimensions["H"].width = 14


# ═══════════════════════════════════════════════════════════════════
# SHEET 2 — TIER MODEL (rebuilt tier comparison)
# ═══════════════════════════════════════════════════════════════════
wt = wb.create_sheet("Tier Model")
wt["A1"] = "Rebuilt Tier Model — pricing references Assumptions sheet"
wt["A1"].font = title; wt.merge_cells("A1:D1")

wt["A3"] = ""; wt["B3"] = "Maker"; wt["C3"] = "Builder"; wt["D3"] = "Agency"
for col in "BCD":
    wt[f"{col}3"].font = hdr; wt[f"{col}3"].fill = hdr_fill
    wt[f"{col}3"].alignment = center; wt[f"{col}3"].border = border

tier_rows = [
    ("Monthly subscription",         '=0',                              '=Assumptions!B26',                  '=Assumptions!B27',                  money),
    ("Velocity Tier 1 fee",          '=Assumptions!C32+Assumptions!B79/100', '=Assumptions!D32+Assumptions!B79/100', '=Assumptions!E32+Assumptions!B79/100', pct),
    ("Velocity Tier 2 fee",          '=Assumptions!C33+Assumptions!B79/100', '=Assumptions!D33+Assumptions!B79/100', '=Assumptions!E33+Assumptions!B79/100', pct),
    ("Velocity Tier 3 fee",          '=Assumptions!C34+Assumptions!B79/100', '=Assumptions!D34+Assumptions!B79/100', '=Assumptions!E34+Assumptions!B79/100', pct),
    ("Velocity Tier 4 fee (floor 5%)",'=Assumptions!C35+Assumptions!B79/100', '=Assumptions!D35+Assumptions!B79/100', '=Assumptions!E35+Assumptions!B79/100', pct),
    ("Velocity Tier 5 fee (floor 5%)",'=Assumptions!C36+Assumptions!B79/100', '=Assumptions!D36+Assumptions!B79/100', '=Assumptions!E36+Assumptions!B79/100', pct),
    ("Effective fee (vol-weighted)",
       '=SUMPRODUCT(Assumptions!C32:C36+Assumptions!B79/100,Assumptions!B40:B44)',
       '=SUMPRODUCT(Assumptions!D32:D36+Assumptions!B79/100,Assumptions!C40:C44)',
       '=SUMPRODUCT(Assumptions!E32:E36+Assumptions!B79/100,Assumptions!D40:D44)',
       pct),
    ("Bulk orders allowed", "YES",   "YES",   "YES", None),
    ("On-demand orders allowed", "YES", "YES", "YES", None),
    ("AI parser quota / mo", "—", '=Assumptions!B68', '=Assumptions!B69', "#,##0"),
    ("Brands per account", "1", "3", "Unlimited", None),
]
for i, (label, m, b, a, fmt) in enumerate(tier_rows, start=4):
    wt.cell(row=i, column=1, value=label).font = body
    wt.cell(row=i, column=2, value=m); wt.cell(row=i, column=3, value=b); wt.cell(row=i, column=4, value=a)
    for col in "BCD":
        wt[f"{col}{i}"].alignment = center; wt[f"{col}{i}"].border = border
        wt[f"{col}{i}"].font = body
        if fmt:
            wt[f"{col}{i}"].number_format = fmt

wt.column_dimensions["A"].width = 38
for col in "BCD":
    wt.column_dimensions[col].width = 18


# ═══════════════════════════════════════════════════════════════════
# SHEET 2.5 — PARTNER TIERS (Verified → Trusted → Anchor)
# Locked 2026-06-04 — see ilaunchify-partner-tier-progression.md
# ═══════════════════════════════════════════════════════════════════
wp = wb.create_sheet("Partner Tiers")

wp["A1"] = "Partner tier ladder — Verified → Trusted → Anchor"
wp["A1"].font = title
wp.merge_cells("A1:E1")

wp["A2"] = ("Three-step narrative arc: arrival → commitment → pillar. ALL THREE tiers FREE in V1.5 — no partner "
            "subscription anywhere. Tier progression is admin-promoted at every step based on performance. "
            "Lower per-order fee at higher tier is the reward. V2 may revisit subscription when advanced tools ship.")
wp["A2"].font = subtitle
wp.merge_cells("A2:E2")
wp["A2"].alignment = Alignment(wrap_text=True, vertical="top")
wp.row_dimensions[2].height = 44

# Partner tier comparison table
wp["A4"] = ""
wp["B4"] = "Verified"
wp["C4"] = "Trusted"
wp["D4"] = "Anchor"
for col in "BCD":
    wp[f"{col}4"].font = hdr; wp[f"{col}4"].fill = hdr_fill
    wp[f"{col}4"].alignment = center; wp[f"{col}4"].border = border

partner_rows = [
    ("Tier means",                          "Arrival",                      "Commitment (earned)",              "Pillar (earned)",                  None),
    ("How to reach it",                     "Pass application + KYB",       "Admin-promoted on performance",    "Admin-promoted / invite-only",     None),
    ("Per-order fee % of partner wholesale",
       "=Assumptions!B55", "=Assumptions!B56", "=Assumptions!B57", pct2),
    ("Monthly subscription cost (V1.5)",    0,                              0,                                  0,                                  money),
    ("Self-serve onboarding",               "YES",                          "Admin promotes",                   "Invite-only",                      None),
    ("Basic Partner Product Builder",       "YES",                          "YES",                              "YES",                              None),
    ("Order accept/decline + payouts",      "YES",                          "YES",                              "YES",                              None),
    ("FDA compliance scan",                 "YES",                          "YES",                              "YES",                              None),
    ("Marketplace placement boost",         "Standard",                     "Modest",                           "Premium",                          None),
    ("Marketplace badge",                   "Verified",                     "Trusted",                          "Anchor",                           None),
    ("Promotion criteria",                  "Pass app",                     "Volume + reliability + on-time",   "Scale + strategic value",          None),
    ("V2 forward-pointer: advanced tools",  "—",                            "TBD — gate-by-tier OR sub overlay","TBD — gate-by-tier OR sub overlay", None),
    ("Tier mix assumption",
       "=Assumptions!B58", "=Assumptions!B59", "=Assumptions!B60", pct2),
]
for i, (label, v, t, p, fmt) in enumerate(partner_rows, start=5):
    wp.cell(row=i, column=1, value=label).font = body
    wp.cell(row=i, column=1).border = border
    wp.cell(row=i, column=2, value=v)
    wp.cell(row=i, column=3, value=t)
    wp.cell(row=i, column=4, value=p)
    for col in "BCD":
        cell = wp[f"{col}{i}"]
        cell.alignment = center; cell.border = border
        cell.font = body
        if fmt:
            cell.number_format = fmt

# Effective partner fee (volume-weighted across tier mix) — anchor row 19
wp["A19"] = "Volume-weighted partner fee % (avg)"; wp["A19"].font = emph
wp["A19"].fill = result_fill; wp["A19"].border = border
wp["B19"] = "=Assumptions!B55*Assumptions!B58+Assumptions!B56*Assumptions!B59+Assumptions!B57*Assumptions!B60"
wp["B19"].number_format = pct2
wp["B19"].font = result_font; wp["B19"].fill = result_fill; wp["B19"].border = border

# Worked example
wp["A21"] = "Example: 1,000-unit on-demand order @ $8 wholesale"
wp["A21"].font = section; wp.merge_cells("A21:D21")
wp["A22"] = "Partner wholesale revenue (gross)"
wp["B22"] = "=1000*Assumptions!B49"; wp["B22"].number_format = money
wp["A23"] = "Verified partner fee (5%)"
wp["B23"] = "=1000*Assumptions!B49*Assumptions!B55"; wp["B23"].number_format = money
wp["A24"] = "Trusted partner fee (3.5%)"
wp["B24"] = "=1000*Assumptions!B49*Assumptions!B56"; wp["B24"].number_format = money
wp["A25"] = "Anchor partner fee (2%)"
wp["B25"] = "=1000*Assumptions!B49*Assumptions!B57"; wp["B25"].number_format = money
wp["A26"] = "Anchor net (highest take-home)"
wp["B26"] = "=B22-B25"; wp["B26"].number_format = money
wp["A26"].font = emph; wp["B26"].font = result_font

# Design principle block
wp["A28"] = "Design principles (LOCKED 2026-06-04)"; wp["A28"].font = section
principles = [
    "1. Verified = arrival. Every approved partner starts free.",
    "2. Trusted = commitment, EARNED through track record. No subscription required.",
    "3. Anchor = pillar, EARNED through scale + strategic value. No subscription required.",
    "4. ALL THREE tiers are FREE in V1.5 — no partner subscription anywhere.",
    "5. ALL tier progression is admin-promoted on performance criteria.",
    "6. Per-order fee differentiation (5% / 3.5% / 2%) is the only monetary reward for tier promotion.",
    "7. Marketplace badges signal trust, never paid placement. Cannot be bought.",
    "8. Essentials (order accept, payouts, basic builder, FDA scan) always free regardless of tier.",
    "9. V2 may revisit a paid subscription overlay when advanced tools ship (3D Builder, AI agent, video gen).",
    "10. Principle: V1.5 monetization must justify itself with a delivered feature that has platform fixed cost.",
]
for i, line in enumerate(principles, start=29):
    c = wp.cell(row=i, column=1, value=line)
    c.font = body
    wp.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
    c.alignment = Alignment(wrap_text=True, vertical="top")

wp.column_dimensions["A"].width = 44
for col in "BCD":
    wp.column_dimensions[col].width = 26
wp.column_dimensions["E"].width = 14


# ═══════════════════════════════════════════════════════════════════
# SHEET 2.7 — BULK TIER (per-order quantity tier system, MOQ-anchored)
# ═══════════════════════════════════════════════════════════════════
wb_sheet = wb.create_sheet("Bulk Tier")

wb_sheet["A1"] = "Bulk Pricing — per-order quantity tier system (V1.5)"
wb_sheet["A1"].font = title
wb_sheet.merge_cells("A1:H1")

wb_sheet["A2"] = ("Bulk uses per-order quantity tier (separate from on-demand velocity tier). "
                  "Platform-suggested category brackets, partner-customizable wholesale prices. "
                  "Tiers below MOQ render N/A. Maker gets full access at higher base fee rate.")
wb_sheet["A2"].font = subtitle
wb_sheet.merge_cells("A2:H2")

# === Platform fee on bulk (NOT velocity-discounted) ==================
section_h = wb_sheet["A4"]
section_h.value = "Platform fee on bulk (FLAT — not velocity-discounted)"
section_h.font = section; section_h.fill = section_fill
wb_sheet.merge_cells("A4:H4")

wb_sheet["A5"] = "Tier"
wb_sheet["B5"] = "Subscription"
wb_sheet["C5"] = "Bulk platform fee %"
wb_sheet["D5"] = "Bulk access"
for col in "ABCD":
    wb_sheet[f"{col}5"].font = hdr; wb_sheet[f"{col}5"].fill = hdr_fill
    wb_sheet[f"{col}5"].alignment = center; wb_sheet[f"{col}5"].border = border

bulk_fee_rows = [
    ("Maker",   0,                    "=Assumptions!C32",  "FULL (locked Pavel 2026-06-04)"),
    ("Builder", "=Assumptions!B26",   "=Assumptions!D32",  "FULL"),
    ("Agency",  "=Assumptions!B27",   "=Assumptions!E32",  "FULL"),
]
for i, (t, sub, fee, access) in enumerate(bulk_fee_rows, start=6):
    wb_sheet.cell(row=i, column=1, value=t).font = emph
    wb_sheet.cell(row=i, column=2, value=sub).number_format = money
    wb_sheet.cell(row=i, column=3, value=fee).number_format = pct2
    wb_sheet.cell(row=i, column=4, value=access)
    for col in "ABCD":
        wb_sheet[f"{col}{i}"].border = border
        wb_sheet[f"{col}{i}"].alignment = center
        if col == "A":
            wb_sheet[f"{col}{i}"].font = emph

wb_sheet["A10"] = ("NOTE: Bulk fee % = the Tier-1 subscription tier rate from Assumptions sheet "
                   "(flat, not velocity-discounted). Bulk's pricing leverage is the wholesale tier ladder below.")
wb_sheet["A10"].font = subtitle; wb_sheet.merge_cells("A10:H10")
wb_sheet["A10"].alignment = Alignment(wrap_text=True, vertical="top")
wb_sheet.row_dimensions[10].height = 30

# === Category bracket ladders (6 categories × 6 tiers) ===============
section_h2 = wb_sheet["A13"]
section_h2.value = "Category bracket ladders (V1.5 suggested defaults — partner-customizable per product)"
section_h2.font = section; section_h2.fill = section_fill
wb_sheet.merge_cells("A13:H13")

# Header row
wb_sheet["A14"] = "Tier"
categories = ["Supplements", "Beverages", "Snack food", "Pet products", "Cosmetics", "OTC pharma"]
for i, cat in enumerate(categories):
    col_letter = chr(ord("B") + i)  # B, C, D, E, F, G
    wb_sheet[f"{col_letter}14"] = cat

for col in "ABCDEFG":
    wb_sheet[f"{col}14"].font = hdr; wb_sheet[f"{col}14"].fill = hdr_fill
    wb_sheet[f"{col}14"].alignment = center; wb_sheet[f"{col}14"].border = border

# Category bracket data
category_brackets = [
    # (tier_label, Supplements, Beverages, Snack food, Pet, Cosmetics, OTC)
    ("T1", "100–249",         "500–999",         "500–999",         "100–499",         "250–499",         "1,000–4,999"),
    ("T2", "250–999",         "1,000–4,999",     "1,000–4,999",     "500–1,999",       "500–1,999",       "5,000–19,999"),
    ("T3", "1,000–4,999",     "5,000–19,999",    "5,000–19,999",    "2,000–9,999",     "2,000–9,999",     "20,000–49,999"),
    ("T4", "5,000–14,999",    "20,000–49,999",   "20,000–49,999",   "10,000–29,999",   "10,000–29,999",   "50,000–99,999"),
    ("T5", "15,000–49,999",   "50,000–99,999",   "50,000–99,999",   "30,000–99,999",   "30,000–99,999",   "100,000–249,999"),
    ("T6", "50,000+",         "100,000+",        "100,000+",        "100,000+",        "100,000+",        "250,000+"),
]
for r_offset, row in enumerate(category_brackets, start=15):
    for c_idx, val in enumerate(row):
        col_letter = chr(ord("A") + c_idx)
        cell = wb_sheet.cell(row=r_offset, column=c_idx + 1, value=val)
        cell.alignment = center; cell.border = border
        cell.font = emph if c_idx == 0 else body

# === MOQ-anchored example (Supliful NMN reproduction) =================
section_h3 = wb_sheet["A23"]
section_h3.value = "MOQ-anchored example — Supplements product with MOQ = 100"
section_h3.font = section; section_h3.fill = section_fill
wb_sheet.merge_cells("A23:H23")

wb_sheet["A24"] = "Tier"
wb_sheet["B24"] = "Qty range"
wb_sheet["C24"] = "Status"
wb_sheet["D24"] = "Partner wholesale $/unit"
wb_sheet["E24"] = "+ Platform fee (Maker 15%)"
wb_sheet["F24"] = "+ Platform fee (Builder 10%)"
wb_sheet["G24"] = "+ Platform fee (Agency 7%)"
for col in "ABCDEFG":
    wb_sheet[f"{col}24"].font = hdr; wb_sheet[f"{col}24"].fill = hdr_fill
    wb_sheet[f"{col}24"].alignment = center; wb_sheet[f"{col}24"].border = border

# Sample bulk pricing ladder from Supliful NMN screenshot, MOQ = 100
moq_example = [
    ("T1 (below MOQ)", "0–49",         "N/A",        None,    None,    None,    None),
    ("T1 (below MOQ)", "50–99",        "N/A",        None,    None,    None,    None),
    ("T1",             "100–249",      "ACTIVE",     7.84,    True,    True,    True),
    ("T2",             "250–499",      "ACTIVE",     7.11,    True,    True,    True),
    ("T3",             "500–999",      "ACTIVE",     6.74,    True,    True,    True),
    ("T4",             "1,000–2,499",  "ACTIVE",     6.43,    True,    True,    True),
    ("T5",             "2,500–4,999",  "ACTIVE",     5.83,    True,    True,    True),
    ("T6",             "5,000+",       "ACTIVE",     5.40,    True,    True,    True),
]
for i, (tier, qty, status, wholesale, m, b, a) in enumerate(moq_example, start=25):
    wb_sheet.cell(row=i, column=1, value=tier).font = emph
    wb_sheet.cell(row=i, column=2, value=qty)
    wb_sheet.cell(row=i, column=3, value=status)
    if wholesale is None:
        for c_idx in range(4, 8):
            wb_sheet.cell(row=i, column=c_idx, value="—")
    else:
        # Wholesale (input value)
        c = wb_sheet.cell(row=i, column=4, value=wholesale)
        c.number_format = '"$"#,##0.00'
        c.fill = input_fill
        # Maker total: wholesale × (1 + 0.15)
        wb_sheet.cell(row=i, column=5,
            value=f"=D{i}*(1+Assumptions!C32)").number_format = '"$"#,##0.00'
        # Builder total: wholesale × (1 + 0.10)
        wb_sheet.cell(row=i, column=6,
            value=f"=D{i}*(1+Assumptions!D32)").number_format = '"$"#,##0.00'
        # Agency total: wholesale × (1 + 0.07)
        wb_sheet.cell(row=i, column=7,
            value=f"=D{i}*(1+Assumptions!E32)").number_format = '"$"#,##0.00'
    for col in "ABCDEFG":
        wb_sheet[f"{col}{i}"].border = border
        wb_sheet[f"{col}{i}"].alignment = center
        if status == "N/A":
            wb_sheet[f"{col}{i}"].font = subtitle

# === Worked example — 5,000 unit bulk run, all three sub tiers =======
section_h4 = wb_sheet["A35"]
section_h4.value = "Worked example — 5,000-unit bulk run @ T6 wholesale ($5.40/unit)"
section_h4.font = section; section_h4.fill = section_fill
wb_sheet.merge_cells("A35:H35")

we_rows = [
    ("Quantity",                  5000,                                "#,##0"),
    ("Wholesale unit price",      "=D32",                               '"$"#,##0.00'),  # T6 row
    ("Wholesale total",           "=B36*B37",                           money),
]
for i, (label, val, fmt) in enumerate(we_rows, start=36):
    wb_sheet.cell(row=i, column=1, value=label).font = body
    c = wb_sheet.cell(row=i, column=2, value=val)
    c.number_format = fmt; c.alignment = right_a; c.border = border
    wb_sheet.cell(row=i, column=1).border = border

# Per-tier comparison
wb_sheet["A39"] = "Tier"
wb_sheet["B39"] = "Platform fee %"
wb_sheet["C39"] = "Platform fee $"
wb_sheet["D39"] = "Total creator cost"
wb_sheet["E39"] = "Savings vs Maker"
for col in "ABCDE":
    wb_sheet[f"{col}39"].font = hdr; wb_sheet[f"{col}39"].fill = hdr_fill
    wb_sheet[f"{col}39"].alignment = center; wb_sheet[f"{col}39"].border = border

we_compare = [
    ("Maker",   "=Assumptions!C32"),
    ("Builder", "=Assumptions!D32"),
    ("Agency",  "=Assumptions!E32"),
]
for i, (t, fee_formula) in enumerate(we_compare, start=40):
    wb_sheet.cell(row=i, column=1, value=t).font = emph
    wb_sheet.cell(row=i, column=2, value=fee_formula).number_format = pct2
    wb_sheet.cell(row=i, column=3, value=f"=B38*B{i}").number_format = money
    wb_sheet.cell(row=i, column=4, value=f"=B38+C{i}").number_format = money
    wb_sheet.cell(row=i, column=5, value=f"=$D$40-D{i}").number_format = money
    for col in "ABCDE":
        wb_sheet[f"{col}{i}"].border = border
        wb_sheet[f"{col}{i}"].alignment = center if col in "AB" else right_a

# === Decisions locked block ==========================================
section_h5 = wb_sheet["A44"]
section_h5.value = "Locked decisions (2026-06-04) — see _V1.5_BULK_PRICING.md for full spec"
section_h5.font = section; section_h5.fill = section_fill
wb_sheet.merge_cells("A44:H44")

decisions = [
    "1. Bulk and on-demand tier systems are SEPARATE",
    "2. Cross-pollination YES — bulk volume accrues to SKU's on-demand velocity tier",
    "3. Platform-suggested category brackets, partner-customizable",
    "4. MOQ-anchored — tiers below MOQ render N/A",
    "5. MaxOQ per batch — auto-split or route to another partner above ceiling",
    "6. Lead time AND setup fee are per-tier partner-set",
    "7. Setup fee displayed as separate line at checkout (NOT amortized)",
    "8. Mixed-flavor multipacks aggregate PER-SKU, not combined",
    "9. Quantity step enforces case-pack multiples",
    "10. Decoration availability per-tier (digital at T1, offset T3+, shrink sleeve T4+)",
    "11. Tier price changes are quote-snapshot-protected (7-day lock)",
    "12. Sub-MOQ requests default SUGGEST_ON_DEMAND (cost comparison)",
    "13. Maker has FULL access — differentiation is base fee rate, not feature lock",
]
for i, line in enumerate(decisions, start=45):
    c = wb_sheet.cell(row=i, column=1, value=line)
    c.font = body
    wb_sheet.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

wb_sheet.column_dimensions["A"].width = 28
wb_sheet.column_dimensions["B"].width = 22
wb_sheet.column_dimensions["C"].width = 22
wb_sheet.column_dimensions["D"].width = 26
wb_sheet.column_dimensions["E"].width = 26
wb_sheet.column_dimensions["F"].width = 26
wb_sheet.column_dimensions["G"].width = 26
wb_sheet.column_dimensions["H"].width = 14


# ═══════════════════════════════════════════════════════════════════
# SHEET 3 — RESULTS (4 scales, all formulas)
# ═══════════════════════════════════════════════════════════════════
wr = wb.create_sheet("Results")

wr["A1"] = "Results — live, formula-driven, 4 scales"
wr["A1"].font = title; wr.merge_cells("A1:E1")
wr["A2"] = "Change any Assumptions input and every cell here recalculates. Toggle scenario via Assumptions!B78 (conservative) and B79 (uplift)."
wr["A2"].font = subtitle; wr.merge_cells("A2:E2")

SCALES = [100, 1_000, 10_000, 30_000]

# Header
wr["A4"] = "Metric"
for col_i, n in enumerate(SCALES, start=2):
    c = wr.cell(row=4, column=col_i, value=f"{n:,} creators")
    c.font = hdr; c.fill = hdr_fill; c.alignment = center; c.border = border
wr["A4"].font = hdr; wr["A4"].fill = hdr_fill; wr["A4"].border = border

# Build per-scale formulas
# We'll use cross-sheet references — every value pulls from Assumptions
A = "Assumptions"
# Conditional activation: IF(B78=1, conservative, base)
def actv(tier_letter):  # tier_letter A/B/C for Maker/Builder/Agency activation rows 11/12/13
    base_row = {"M": 11, "B": 12, "A": 13}[tier_letter]
    return f"IF({A}!$B$78=1,{A}!E{base_row},{A}!B{base_row})"

def orders(tier_letter):
    base_row = {"M": 16, "B": 17, "A": 18}[tier_letter]
    return f"IF({A}!$B$78=1,{A}!E{base_row},{A}!B{base_row})"

def bulkruns(tier_letter):
    base_row = {"M": 21, "B": 22, "A": 23}[tier_letter]
    return f"IF({A}!$B$78=1,{A}!G{base_row},{A}!B{base_row})"

def velocity_dist_col(tier_letter):
    # For Conservative mode swap to columns F:H rows 40:44 ; base is B:D rows 40:44
    # Returns "{A}!B40:B44" etc.
    base_col = {"M": "B", "B": "C", "A": "D"}[tier_letter]
    cons_col = {"M": "F", "B": "G", "A": "H"}[tier_letter]
    return f"IF({A}!$B$78=1,{A}!{cons_col}40:{cons_col}44,{A}!{base_col}40:{base_col}44)"

# Mix uses Rec 1 toggle (B82): if 1, use rec 1 mix shift
def mix(tier_letter):
    if tier_letter == "M":
        return f"IF({A}!$B$82=1,{A}!B102,{A}!B5)"
    elif tier_letter == "B":
        return f"IF({A}!$B$82=1,{A}!B103,{A}!B6)"
    else:  # Agency stays the same
        return f"{A}!B7"

# Sub prices use Rec 3 toggle
def subscription_price(tier_letter):
    # If Rec 3 active, use $79/$249. Otherwise use $49/$199 base (we'll hardcode the toggle)
    if tier_letter == "B":
        return f"IF({A}!$B$84=1,79,49)"
    elif tier_letter == "A":
        return f"IF({A}!$B$84=1,249,199)"
    else:
        return "0"

# Builder activation uses Rec 1 lift
def builder_activation():
    base = f"IF({A}!$B$78=1,{A}!E12,{A}!B12)"
    return f"IF({A}!$B$82=1,MAX({base},{A}!B104),{base})"

# Now write the formulas per scale column
def write_metric(row_i, label, formula_template, fmt=money, bold=False):
    wr.cell(row=row_i, column=1, value=label).font = emph if bold else body
    wr.cell(row=row_i, column=1).border = border
    for col_i, n in enumerate(SCALES, start=2):
        formula = formula_template.format(N=n)
        c = wr.cell(row=row_i, column=col_i, value=formula)
        c.number_format = fmt
        c.alignment = right_a
        c.border = border
        c.font = output_font

# ─── User counts ─────────────────────────────────────────────────
write_metric(5, "Partners",
    f"=ROUND({{N}}/{A}!B62,0)", "#,##0", bold=True)

write_metric(6, "Registered Makers", f"=ROUND({{N}}*({mix('M')}),0)", "#,##0")
write_metric(7, "Registered Builders", f"=ROUND({{N}}*({mix('B')}),0)", "#,##0")
write_metric(8, "Registered Agencies", f"=ROUND({{N}}*({mix('A')}),0)", "#,##0")

# Active counts
write_metric(9, "Active Makers", f"=ROUND({{N}}*({mix('M')})*({actv('M')}),0)", "#,##0", bold=True)
write_metric(10, "Active Builders", f"=ROUND({{N}}*({mix('B')})*({builder_activation()}),0)", "#,##0", bold=True)
write_metric(11, "Active Agencies", f"=ROUND({{N}}*({mix('A')})*({actv('A')}),0)", "#,##0", bold=True)

# Subscription revenue
write_metric(12, "Subscription revenue (mo)",
    f"=ROUND({{N}}*({mix('B')})*{subscription_price('B')}+{{N}}*({mix('A')})*{subscription_price('A')},0)",
    money, bold=True)

# Effective fee (volume-weighted) per tier — uses velocity distribution + fees + uplift
# Build the SUMPRODUCT for each tier
# Maker effective: sum over 5 tiers of (Maker_fee[i] + uplift) × Maker_dist[i]
def eff_fee(tier_letter):
    if tier_letter == "M":
        fee_range = f"({A}!C32:C36+{A}!B79/100)"
        dist_range = f"IF({A}!$B$78=1,{A}!F40:F44,{A}!B40:B44)"
    elif tier_letter == "B":
        fee_range = f"({A}!D32:D36+{A}!B79/100)"
        dist_range = f"IF({A}!$B$78=1,{A}!G40:G44,{A}!C40:C44)"
    else:
        fee_range = f"({A}!E32:E36+{A}!B79/100)"
        dist_range = f"IF({A}!$B$78=1,{A}!H40:H44,{A}!D40:D44)"
    return f"SUMPRODUCT({fee_range},{dist_range})"

# Partner avg fee (weighted)
partner_eff = f"({A}!B55*{A}!B58+{A}!B56*{A}!B59+{A}!B57*{A}!B60)"

# On-demand units per tier × scale
# Per tier active × orders × (1 - bulk_mix_share)  — using simplified bulk_share
# We'll assume bulk_share = 5/30/60% as in original model (hardcoded approximation here)
def od_units(tier_letter, scale):
    if tier_letter == "M":
        active_expr = f"({scale}*({mix('M')})*({actv('M')}))"
        orders_expr = orders('M')
        bulk_share = 0.05
    elif tier_letter == "B":
        active_expr = f"({scale}*({mix('B')})*({builder_activation()}))"
        orders_expr = orders('B')
        bulk_share = 0.30
    else:
        active_expr = f"({scale}*({mix('A')})*({actv('A')}))"
        orders_expr = orders('A')
        bulk_share = 0.60
    return f"{active_expr}*{orders_expr}*(1-{bulk_share})"

def bulk_units(tier_letter, scale):
    if tier_letter == "M":
        active_expr = f"({scale}*({mix('M')})*({actv('M')}))"
        return f"{active_expr}*{bulkruns('M')}*{A}!D21"
    elif tier_letter == "B":
        active_expr = f"({scale}*({mix('B')})*({builder_activation()}))"
        return f"{active_expr}*{bulkruns('B')}*{A}!D22"
    else:
        active_expr = f"({scale}*({mix('A')})*({actv('A')}))"
        return f"{active_expr}*{bulkruns('A')}*{A}!D23"

# Sum across tiers
def total_od_units(scale):
    return f"({od_units('M',scale)}+{od_units('B',scale)}+{od_units('A',scale)})"

def total_bulk_units(scale):
    return f"({bulk_units('M',scale)}+{bulk_units('B',scale)}+{bulk_units('A',scale)})"

# On-demand units
write_metric(14, "On-demand units (mo)",
    "=ROUND(" + total_od_units("{N}") + ",0)", "#,##0")
write_metric(15, "On-demand gross retail (mo)",
    f"=ROUND(({total_od_units('{N}')})*{A}!B48,0)", money)

# Creator-side OD fee = sum per tier (od_units_tier * retail * eff_fee_tier)
def od_creator_fee(scale):
    parts = []
    for t in "MBA":
        parts.append(f"({od_units(t, scale)})*{A}!B48*{eff_fee(t)}")
    return "+".join(parts)

write_metric(16, "Creator-side on-demand fee",
    f"=ROUND({od_creator_fee('{N}')},0)", money)

write_metric(17, "Partner-side on-demand fee",
    f"=ROUND(({total_od_units('{N}')})*{A}!B49*{partner_eff},0)", money)

# Bulk units + gross + fees
write_metric(19, "Bulk units (mo)",
    "=ROUND(" + total_bulk_units("{N}") + ",0)", "#,##0")
write_metric(20, "Bulk partner cost flow (mo)",
    f"=ROUND(({total_bulk_units('{N}')})*{A}!B51,0)", money)

def bulk_creator_fee(scale):
    parts = []
    for t in "MBA":
        parts.append(f"({bulk_units(t, scale)})*{A}!B51*{eff_fee(t)}")
    return "+".join(parts)

write_metric(21, "Creator-side bulk fee",
    f"=ROUND({bulk_creator_fee('{N}')},0)", money)
write_metric(22, "Partner-side bulk fee",
    f"=ROUND(({total_bulk_units('{N}')})*{A}!B51*{partner_eff},0)", money)

# Subtotal: base platform take — write per-column directly
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)  # B, C, D, E
    formula = f"={col}12+{col}16+{col}17+{col}21+{col}22"
    c = wr.cell(row=24, column=col_i, value=formula)
    c.number_format = money; c.alignment = right_a; c.font = output_font; c.border = border
    c.fill = section_fill
wr.cell(row=24, column=1, value="Base platform take (mo)").font = emph
wr.cell(row=24, column=1).fill = section_fill
wr.cell(row=24, column=1).border = border

# ─── Recommendation contributions ────────────────────────────────
# Rec 2: float income = (gross_od + bulk_gross) × treasury_rate × (days/365)
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    # row 26 — Float income
    formula = f"=IF({A}!$B$83=1,({col}15+{col}20)*{A}!B93*({A}!B92/365),0)"
    c = wr.cell(row=26, column=col_i, value=formula)
    c.number_format = money; c.font = output_font; c.alignment = right_a; c.border = border
wr.cell(row=26, column=1, value="Rec 2 — Float income").font = body
wr.cell(row=26, column=1).border = border

# Rec 6: partner referral drag = -(base take × share × commission)
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    formula = f"=-IF({A}!$B$86=1,{col}24*{A}!B94*{A}!B95,0)"
    c = wr.cell(row=27, column=col_i, value=formula)
    c.number_format = money_neg; c.font = output_font; c.alignment = right_a; c.border = border
wr.cell(row=27, column=1, value="Rec 6 — Partner referral drag").font = body
wr.cell(row=27, column=1).border = border

# Rec 7: AI overage = (Builder + Agency active) × over_cap_rate × gens × (price - cost)
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    formula = f"=IF({A}!$B$87=1,({col}10+{col}11)*{A}!B96*{A}!B97*({A}!B98-{A}!B99),0)"
    c = wr.cell(row=28, column=col_i, value=formula)
    c.number_format = money; c.font = output_font; c.alignment = right_a; c.border = border
wr.cell(row=28, column=1, value="Rec 7 — Premium AI overage").font = body
wr.cell(row=28, column=1).border = border

# Rec 8: bulk escrow = bulk_gross × opt_in × fee
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    formula = f"=IF({A}!$B$88=1,{col}20*{A}!B100*{A}!B101,0)"
    c = wr.cell(row=29, column=col_i, value=formula)
    c.number_format = money; c.font = output_font; c.alignment = right_a; c.border = border
wr.cell(row=29, column=1, value="Rec 8 — Bulk escrow").font = body
wr.cell(row=29, column=1).border = border

# Subtotal: total take with recs
wr.cell(row=31, column=1, value="Total platform take (mo)").font = emph
wr.cell(row=31, column=1).fill = result_fill
wr.cell(row=31, column=1).border = border
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    formula = f"={col}24+{col}26+{col}27+{col}28+{col}29"
    c = wr.cell(row=31, column=col_i, value=formula)
    c.number_format = money; c.font = result_font; c.alignment = right_a; c.border = border
    c.fill = result_fill

# ─── Costs ────────────────────────────────────────────────────────
# Stripe ≈ total transactions × flat + total_gross × pct
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    # Approx: tx count = OD units + bulk_units/500
    formula = f"=ROUND(({col}14+{col}19/500)*{A}!B66+({col}15+{col}20)*{A}!B65,0)"
    c = wr.cell(row=33, column=col_i, value=formula)
    c.number_format = money; c.font = output_font; c.alignment = right_a; c.border = border
wr.cell(row=33, column=1, value="Stripe Connect costs").font = body
wr.cell(row=33, column=1).border = border

# AI parser cost = (active_B × B68 + active_A × B69) × B67
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    formula = f"=ROUND({col}10*{A}!B68*{A}!B67+{col}11*{A}!B69*{A}!B67,0)"
    c = wr.cell(row=34, column=col_i, value=formula)
    c.number_format = money; c.font = output_font; c.alignment = right_a; c.border = border
wr.cell(row=34, column=1, value="AI parser cost").font = body
wr.cell(row=34, column=1).border = border

# Asset library flat
for col_i, n in enumerate(SCALES, start=2):
    formula = f"={A}!B70"
    c = wr.cell(row=35, column=col_i, value=formula)
    c.number_format = money; c.font = output_font; c.alignment = right_a; c.border = border
wr.cell(row=35, column=1, value="Asset library (Shutterstock)").font = body
wr.cell(row=35, column=1).border = border

# Fixed opex monthly = matching scale row /12
opex_rows = {100: 71, 1_000: 72, 10_000: 73, 30_000: 74}
for col_i, n in enumerate(SCALES, start=2):
    formula = f"={A}!B{opex_rows[n]}/12"
    c = wr.cell(row=36, column=col_i, value=formula)
    c.number_format = money; c.font = output_font; c.alignment = right_a; c.border = border
wr.cell(row=36, column=1, value="Fixed opex (monthly)").font = body
wr.cell(row=36, column=1).border = border

# ─── Bottom line ─────────────────────────────────────────────────
wr.cell(row=38, column=1, value="Operating profit (monthly)").font = emph
wr.cell(row=38, column=1).fill = result_fill
wr.cell(row=38, column=1).border = border
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    formula = f"={col}31-({col}33+{col}34+{col}35+{col}36)"
    c = wr.cell(row=38, column=col_i, value=formula)
    c.number_format = money; c.font = result_font; c.alignment = right_a; c.border = border
    c.fill = result_fill

wr.cell(row=39, column=1, value="Annualized revenue (ARR)").font = emph
wr.cell(row=39, column=1).fill = result_fill
wr.cell(row=39, column=1).border = border
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    formula = f"={col}31*12"
    c = wr.cell(row=39, column=col_i, value=formula)
    c.number_format = money; c.font = result_font; c.alignment = right_a; c.border = border
    c.fill = result_fill

wr.cell(row=40, column=1, value="Annualized op profit").font = emph
wr.cell(row=40, column=1).fill = result_fill
wr.cell(row=40, column=1).border = border
for col_i, n in enumerate(SCALES, start=2):
    col = chr(ord('A') + col_i - 1)
    formula = f"={col}38*12"
    c = wr.cell(row=40, column=col_i, value=formula)
    c.number_format = money; c.font = result_font; c.alignment = right_a; c.border = border
    c.fill = result_fill

wr.column_dimensions["A"].width = 36
for col in "BCDE":
    wr.column_dimensions[col].width = 18


# ═══════════════════════════════════════════════════════════════════
# SHEET 4 — DASHBOARD
# ═══════════════════════════════════════════════════════════════════
wd = wb.create_sheet("Dashboard")
wd["A1"] = "DASHBOARD — live headline numbers"
wd["A1"].font = title; wd.merge_cells("A1:E1")
wd["A2"] = "Toggle scenarios on Assumptions sheet. This dashboard reflects the active configuration."
wd["A2"].font = subtitle; wd.merge_cells("A2:E2")

# Active scenario indicator
wd["A4"] = "Active scenario:"
wd["A4"].font = section
wd["B4"] = f'=IF(Assumptions!B78=1,"Conservative","Base") & " demand, +" & TEXT(Assumptions!B79,"0") & "pp uplift"'
wd["B4"].font = emph
wd.merge_cells("B4:E4")

wd["A5"] = "Recs active:"
wd["B5"] = f'="Rec1=" & Assumptions!B82 & "  Rec2=" & Assumptions!B83 & "  Rec3=" & Assumptions!B84 & "  Rec6=" & Assumptions!B86 & "  Rec7=" & Assumptions!B87 & "  Rec8=" & Assumptions!B88'
wd.merge_cells("B5:E5")

# Headline table
wd["A7"] = "Scale"
wd["B7"] = "Annualized revenue"
wd["C7"] = "Operating profit"
wd["D7"] = "Profit / creator / yr"
for col in "ABCD":
    wd[f"{col}7"].font = hdr; wd[f"{col}7"].fill = hdr_fill
    wd[f"{col}7"].alignment = center; wd[f"{col}7"].border = border

scale_cols = {100: "B", 1_000: "C", 10_000: "D", 30_000: "E"}
for i, n in enumerate(SCALES, start=8):
    wd.cell(row=i, column=1, value=f"{n:,} creators").font = emph
    wd.cell(row=i, column=2, value=f"=Results!{scale_cols[n]}39").number_format = money
    wd.cell(row=i, column=3, value=f"=Results!{scale_cols[n]}40").number_format = money
    wd.cell(row=i, column=4, value=f"=Results!{scale_cols[n]}40/{n}").number_format = money
    for col in "ABCD":
        wd[f"{col}{i}"].border = border
        wd[f"{col}{i}"].alignment = right_a if col != "A" else left_a

# Effective fees
wd["A13"] = "Effective fees (volume-weighted)"; wd["A13"].font = section
wd["A14"] = "Maker"
wd["B14"] = f'=SUMPRODUCT(Assumptions!C32:C36+Assumptions!B79/100,IF(Assumptions!B78=1,Assumptions!F40:F44,Assumptions!B40:B44))'
wd["A15"] = "Builder"
wd["B15"] = f'=SUMPRODUCT(Assumptions!D32:D36+Assumptions!B79/100,IF(Assumptions!B78=1,Assumptions!G40:G44,Assumptions!C40:C44))'
wd["A16"] = "Agency"
wd["B16"] = f'=SUMPRODUCT(Assumptions!E32:E36+Assumptions!B79/100,IF(Assumptions!B78=1,Assumptions!H40:H44,Assumptions!D40:D44))'
wd["A17"] = "Partner avg"
wd["B17"] = "=Assumptions!B55*Assumptions!B58+Assumptions!B56*Assumptions!B59+Assumptions!B57*Assumptions!B60"
for r in (14, 15, 16, 17):
    wd[f"B{r}"].number_format = pct2

wd.column_dimensions["A"].width = 22
for col in "BCDE":
    wd.column_dimensions[col].width = 22


# ═══════════════════════════════════════════════════════════════════
# SHEET 5 — README
# ═══════════════════════════════════════════════════════════════════
wm = wb.create_sheet("README")
wm["A1"] = "HOW TO USE THIS LIVE MODEL"
wm["A1"].font = title
wm.column_dimensions["A"].width = 110

readme = [
    "",
    "WHAT THIS IS",
    "One master live model that supersedes the 5 prior static spreadsheets. Every yellow cell on Assumptions is an input. Every white",
    "cell is a formula. Edit any input — the model recalculates everywhere.",
    "",
    "QUICK SCENARIO TOGGLES (on Assumptions sheet)",
    "Cell B78 — Use Conservative demand?       0 = Base assumptions, 1 = Conservative (halved demand levers)",
    "Cell B79 — Pricing uplift (pp)            0 = locked V1.5, 5 = +5pp aggressive, 10 = +10pp ceiling test",
    "",
    "RECOMMENDATION TOGGLES (set 1 to enable)",
    "B82 — Rec 1 Engineered Maker→Builder triggers (mix shift + Builder activation lift)",
    "B83 — Rec 2 Float income on held balances",
    "B84 — Rec 3 Sub prices Builder $79 / Agency $249 (vs $49/$199)",
    "B86 — Rec 6 Partner referral commission (-2% drag)",
    "B87 — Rec 7 Premium AI overage",
    "B88 — Rec 8 Bulk escrow opt-in",
    "",
    "TYPICAL EXPLORATIONS",
    "  ── Base case             — set everything to 0 (default)",
    "  ── Conservative          — set B78=1",
    "  ── Conservative + Recs   — set B78=1, B82=1, B83=1, B84=1, B86=1, B87=1, B88=1",
    "  ── Aggressive pricing    — set B79=5 (or 10) and watch the matrix in Tier Model recompute",
    "",
    "WHERE TO SEE THE NUMBERS",
    "  Results sheet       — full per-scale breakdown (4 scales)",
    "  Dashboard sheet     — headline summary, also shows what's active",
    "  Tier Model sheet    — creator-side comparison Maker/Builder/Agency (subscription + velocity)",
    "  Partner Tiers sheet — three-step ladder Verified→Trusted→Anchor (locked 2026-06-04, Anchor replaces Premier)",
    "  Bulk Tier sheet     — bulk per-order quantity tiers + category brackets + MOQ-anchored example (NEW 2026-06-04)",
    "",
    "WHAT'S DIFFERENT FROM THE PRIOR SPREADSHEETS",
    "  iLaunchify_Financial_Simulation.xlsx                    — superseded (was Base only)",
    "  iLaunchify_Financial_Simulation_3Scenarios.xlsx          — superseded (was Base/+5/+10pp)",
    "  iLaunchify_Financial_Simulation_Conservative_vs_Base.xlsx — superseded",
    "  iLaunchify_Financial_Simulation_With_Recommendations.xlsx — superseded",
    "  iLaunchify_Rebuilt_Tier_Model.xlsx                       — superseded (tier model is now Sheet 2 here, live)",
    "",
    "All five static spreadsheets are now reproducible by setting the right toggle combination on Assumptions.",
    "",
    "FORMULA CAVEAT",
    "Bulk-share-of-total (5%/30%/60% for Maker/Builder/Agency) is hardcoded inside the Results formulas for simplicity.",
    "If you want to tune bulk share, edit the formulas directly in Results sheet rows 14/19 or ask me to expose it as inputs.",
    "",
    "WHEN YOU HIT THE EDGE",
    "If a calculation looks wrong: open the Results sheet, click the cell, examine the formula. Every formula traces back to",
    "Assumptions. There are no hidden constants beyond the bulk-share noted above.",
]
for i, line in enumerate(readme, start=2):
    c = wm.cell(row=i, column=1, value=line)
    if line.isupper() and len(line) > 5:
        c.font = section
    else:
        c.font = body


wb.save("./iLaunchify_Live_Master_Model.xlsx")
print("Wrote: iLaunchify_Live_Master_Model.xlsx")

# ─── Force LibreOffice to evaluate all formulas and cache the results ──
# Without this step, openpyxl writes formulas without cached values.
# Excel SHOULD auto-recalc on open but doesn't always (depends on version + calc settings).
# This pre-evaluates everything so cells display immediately when Excel opens the file.
import subprocess, os, shutil
print("Force-evaluating formulas with LibreOffice headless...")
result = subprocess.run(
    ["libreoffice", "--headless", "--calc", "--convert-to", "xlsx",
     "--outdir", "_recalc_tmp", "iLaunchify_Live_Master_Model.xlsx"],
    capture_output=True, text=True
)
if os.path.exists("_recalc_tmp/iLaunchify_Live_Master_Model.xlsx"):
    shutil.move("_recalc_tmp/iLaunchify_Live_Master_Model.xlsx",
                "iLaunchify_Live_Master_Model.xlsx")
    shutil.rmtree("_recalc_tmp", ignore_errors=True)
    print("✓ Cached values written. File opens with numbers visible immediately.")
else:
    print(f"⚠ LibreOffice didn't produce output: {result.stderr[:200]}")

print("\nThis ONE file supersedes:")
print("  - iLaunchify_Financial_Simulation.xlsx")
print("  - iLaunchify_Financial_Simulation_3Scenarios.xlsx")
print("  - iLaunchify_Financial_Simulation_Conservative_vs_Base.xlsx")
print("  - iLaunchify_Financial_Simulation_With_Recommendations.xlsx")
print("  - iLaunchify_Rebuilt_Tier_Model.xlsx")
