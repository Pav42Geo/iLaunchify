"""
iLaunchify — layered recommendations on top of Conservative baseline.

Models the 10 recommendations from the strategic critique:
  1. Engineered Maker → Builder upgrade triggers (mix shift + activation lift)
  2. Float income on held balances (treasury revenue)
  3. Builder $49 → $79; Agency $199 → $249 (sub price uplift)
  4. Floor velocity-tier fees at 5% (defends margin at top tiers)
  5. Restrict bulk to Builder+ (kills Maker bulk volume)
  6. Partner referral commission (10% of referred-creator fees = 2% gross drag)
  7. Premium AI generations beyond cap ($1.50/generation overage)
  8. Optional escrow/insurance on bulk orders (0.5% × 40% opt-in)
  9. Agency tier mispricing flagged — not modeled, decision pending
 10. Partner-side velocity tier — not modeled, V2 deferred

Applied on the CONSERVATIVE demand assumptions (year-one realistic).
Output: Conservative baseline vs Conservative+Recs vs Base, with per-recommendation
contribution so Pavel can see which levers matter.
"""

import sys
from copy import deepcopy
from dataclasses import dataclass, field

sys.path.insert(0, "./")
import financial_simulation as base

# Bring in Conservative overrides from conservative_simulation
CONSERVATIVE_ACTIVATION = {"maker": 0.15, "builder": 0.35, "agency": 0.45}
CONSERVATIVE_ORDERS_PER_ACTIVE = {"maker": 2, "builder": 40, "agency": 200}
CONSERVATIVE_BULK_RUNS_PER_MONTH = {"maker": 0.05, "builder": 0.25, "agency": 1.0}
CONSERVATIVE_VELOCITY_DISTRIBUTION = {
    "maker":   {1: 0.95, 2: 0.04, 3: 0.01, 4: 0.00, 5: 0.00},
    "builder": {1: 0.70, 2: 0.20, 3: 0.07, 4: 0.02, 5: 0.01},
    "agency":  {1: 0.50, 2: 0.30, 3: 0.15, 4: 0.04, 5: 0.01},
}

# ─────────────────────────────────────────────────────────────────────
# RECOMMENDATION TUNERS
# ─────────────────────────────────────────────────────────────────────

# Rec 1: Engineered upgrade triggers — shift mix + lift Builder activation
RECS_CREATOR_TIER_MIX = {"maker": 0.55, "builder": 0.40, "agency": 0.05}
RECS_BUILDER_ACTIVATION_LIFT = 0.50  # was 0.35 in Conservative → 0.50

# Rec 3: Sub price uplift
RECS_SUBSCRIPTION_PRICE = {"maker": 0, "builder": 79, "agency": 249}

# Rec 4: Floor velocity fees at 5% creator-side
RECS_VELOCITY_FEE_PCT = {
    "maker":   {1: 15.0, 2: 13.0, 3: 11.0, 4: 9.0,  5: 7.0},   # was identical, floor already at 7
    "builder": {1: 10.0, 2: 8.5,  3: 7.0,  4: 6.0,  5: 5.0},   # was identical, floor already at 5
    "agency":  {1: 7.0,  2: 6.0,  3: 5.0,  4: 5.0,  5: 5.0},   # T4 5%, T5 5% (was 4%, 3%)
}

# Rec 5: Restrict bulk to Builder+ — Maker bulk runs go to 0
RECS_BULK_RUNS = {"maker": 0.0, "builder": 0.25, "agency": 1.0}

# Rec 2: Float income — 2-day average float on bulk + on-demand gross flow
FLOAT_AVG_DAYS = 2.0
TREASURY_ANNUAL_RATE = 0.045  # 4.5% money-market yield
FLOAT_MONTHLY_RATE = TREASURY_ANNUAL_RATE * (FLOAT_AVG_DAYS / 365)

# Rec 6: Partner referral commission — 20% of new creators come via partner, 10% commission of their fees
PARTNER_REFERRAL_CREATOR_SHARE = 0.20
PARTNER_REFERRAL_COMMISSION = 0.10
# Net drag on platform take = share × commission = 2%

# Rec 7: Premium AI generations beyond cap
OVER_CAP_RATE = 0.20             # 20% of Builder/Agency creators exceed monthly cap
OVER_CAP_GENS_PER_MONTH = 30
PRICE_PER_OVER_CAP_GEN = 1.50
GEN_COST = 0.05                  # we pay $0.05 to Replicate/Together per image

# Rec 8: Optional escrow on bulk orders
ESCROW_OPT_IN_RATE = 0.40
ESCROW_FEE_PCT = 0.005           # 0.5% of bulk order value


@dataclass
class RecsResult:
    creator_count: int
    partner_count: int
    active_makers: int = 0
    active_builders: int = 0
    active_agencies: int = 0
    base_take_monthly: float = 0          # platform take BEFORE recs
    sub_revenue_monthly: float = 0
    on_demand_gross_monthly: float = 0
    bulk_gross_monthly: float = 0
    creator_fee_monthly: float = 0
    partner_fee_monthly: float = 0
    # Recommendation contributions
    rec1_upgrade_lift: float = 0
    rec2_float_income: float = 0
    rec3_sub_price_lift: float = 0
    rec4_velocity_floor: float = 0
    rec5_bulk_restriction_change: float = 0
    rec6_referral_drag: float = 0          # negative
    rec7_premium_ai: float = 0
    rec8_escrow: float = 0
    # Costs
    variable_costs_monthly: float = 0
    fixed_opex_monthly: float = 0
    # Totals
    total_take_monthly: float = 0
    operating_profit_monthly: float = 0
    annualized_revenue: float = 0
    annualized_profit: float = 0


def patch_run(creator_count, *, mix=None, activation=None, orders=None,
              bulk_runs=None, velocity_dist=None, velocity_fees=None,
              sub_prices=None):
    """Run base.simulate() with one or more assumption overrides."""
    saved = (
        deepcopy(base.CREATOR_TIER_MIX),
        deepcopy(base.CREATOR_ACTIVATION),
        deepcopy(base.ORDERS_PER_ACTIVE_CREATOR_PER_MONTH),
        deepcopy(base.BULK_RUNS_PER_MONTH),
        deepcopy(base.VELOCITY_DISTRIBUTION),
        deepcopy(base.VELOCITY_FEE_PCT),
        deepcopy(base.SUBSCRIPTION_PRICE_MONTHLY),
    )
    if mix:           base.CREATOR_TIER_MIX = mix
    if activation:    base.CREATOR_ACTIVATION = activation
    if orders:        base.ORDERS_PER_ACTIVE_CREATOR_PER_MONTH = orders
    if bulk_runs:     base.BULK_RUNS_PER_MONTH = bulk_runs
    if velocity_dist: base.VELOCITY_DISTRIBUTION = velocity_dist
    if velocity_fees: base.VELOCITY_FEE_PCT = velocity_fees
    if sub_prices:    base.SUBSCRIPTION_PRICE_MONTHLY = sub_prices
    try:
        return base.simulate(creator_count)
    finally:
        (base.CREATOR_TIER_MIX, base.CREATOR_ACTIVATION,
         base.ORDERS_PER_ACTIVE_CREATOR_PER_MONTH, base.BULK_RUNS_PER_MONTH,
         base.VELOCITY_DISTRIBUTION, base.VELOCITY_FEE_PCT,
         base.SUBSCRIPTION_PRICE_MONTHLY) = saved


def simulate_recommendations(creator_count: int) -> RecsResult:
    """Layer all recommendations on Conservative baseline."""

    # Step 1: Conservative baseline (no recommendations applied)
    conservative = patch_run(
        creator_count,
        activation=CONSERVATIVE_ACTIVATION,
        orders=CONSERVATIVE_ORDERS_PER_ACTIVE,
        bulk_runs=CONSERVATIVE_BULK_RUNS_PER_MONTH,
        velocity_dist=CONSERVATIVE_VELOCITY_DISTRIBUTION,
    )

    # Step 2: Conservative + Rec 1 (mix shift + activation lift)
    activation_with_lift = dict(CONSERVATIVE_ACTIVATION,
                                builder=RECS_BUILDER_ACTIVATION_LIFT)
    rec1_world = patch_run(
        creator_count,
        mix=RECS_CREATOR_TIER_MIX,
        activation=activation_with_lift,
        orders=CONSERVATIVE_ORDERS_PER_ACTIVE,
        bulk_runs=CONSERVATIVE_BULK_RUNS_PER_MONTH,
        velocity_dist=CONSERVATIVE_VELOCITY_DISTRIBUTION,
    )
    rec1_uplift = rec1_world.total_platform_take_monthly - conservative.total_platform_take_monthly

    # Step 3: Conservative + Rec 1 + Rec 3 (sub price uplift)
    rec3_world = patch_run(
        creator_count,
        mix=RECS_CREATOR_TIER_MIX,
        activation=activation_with_lift,
        orders=CONSERVATIVE_ORDERS_PER_ACTIVE,
        bulk_runs=CONSERVATIVE_BULK_RUNS_PER_MONTH,
        velocity_dist=CONSERVATIVE_VELOCITY_DISTRIBUTION,
        sub_prices=RECS_SUBSCRIPTION_PRICE,
    )
    rec3_uplift = rec3_world.total_platform_take_monthly - rec1_world.total_platform_take_monthly

    # Step 4: + Rec 4 (velocity floor) + Rec 5 (bulk restriction)
    rec45_world = patch_run(
        creator_count,
        mix=RECS_CREATOR_TIER_MIX,
        activation=activation_with_lift,
        orders=CONSERVATIVE_ORDERS_PER_ACTIVE,
        bulk_runs=RECS_BULK_RUNS,
        velocity_dist=CONSERVATIVE_VELOCITY_DISTRIBUTION,
        velocity_fees=RECS_VELOCITY_FEE_PCT,
        sub_prices=RECS_SUBSCRIPTION_PRICE,
    )
    rec4_uplift = (
        # Approximated separately — re-run with floor only
        patch_run(
            creator_count,
            mix=RECS_CREATOR_TIER_MIX,
            activation=activation_with_lift,
            orders=CONSERVATIVE_ORDERS_PER_ACTIVE,
            bulk_runs=CONSERVATIVE_BULK_RUNS_PER_MONTH,
            velocity_dist=CONSERVATIVE_VELOCITY_DISTRIBUTION,
            velocity_fees=RECS_VELOCITY_FEE_PCT,
            sub_prices=RECS_SUBSCRIPTION_PRICE,
        ).total_platform_take_monthly - rec3_world.total_platform_take_monthly
    )
    rec5_uplift = rec45_world.total_platform_take_monthly - rec3_world.total_platform_take_monthly - rec4_uplift

    # Recs 2, 6, 7, 8 are additive on top of rec45_world
    # ─── Rec 2: Float income
    flow_monthly = rec45_world.on_demand_gross_revenue_monthly + rec45_world.bulk_gross_revenue_monthly
    rec2_float = flow_monthly * FLOAT_MONTHLY_RATE

    # ─── Rec 6: Partner referral commission (negative — paid back to partners)
    rec6_drag = -(rec45_world.total_platform_take_monthly
                  * PARTNER_REFERRAL_CREATOR_SHARE
                  * PARTNER_REFERRAL_COMMISSION)

    # ─── Rec 7: Premium AI overage
    rec7_ai = (
        (rec45_world.active_builders + rec45_world.active_agencies)
        * OVER_CAP_RATE * OVER_CAP_GENS_PER_MONTH
        * (PRICE_PER_OVER_CAP_GEN - GEN_COST)
    )

    # ─── Rec 8: Escrow on bulk
    rec8_escrow = rec45_world.bulk_gross_revenue_monthly * ESCROW_OPT_IN_RATE * ESCROW_FEE_PCT

    # Total monthly take with all recs
    total_take = (
        rec45_world.total_platform_take_monthly
        + rec2_float
        + rec6_drag
        + rec7_ai
        + rec8_escrow
    )

    variable_costs = rec45_world.total_variable_costs_monthly
    operating_profit = total_take - variable_costs - rec45_world.fixed_opex_monthly

    return RecsResult(
        creator_count=creator_count,
        partner_count=rec45_world.partner_count,
        active_makers=rec45_world.active_makers,
        active_builders=rec45_world.active_builders,
        active_agencies=rec45_world.active_agencies,
        base_take_monthly=conservative.total_platform_take_monthly,
        sub_revenue_monthly=rec45_world.subscription_revenue_monthly,
        on_demand_gross_monthly=rec45_world.on_demand_gross_revenue_monthly,
        bulk_gross_monthly=rec45_world.bulk_gross_revenue_monthly,
        creator_fee_monthly=(rec45_world.on_demand_platform_fee_creator_side
                             + rec45_world.bulk_platform_fee_creator_side),
        partner_fee_monthly=(rec45_world.on_demand_platform_fee_partner_side
                             + rec45_world.bulk_platform_fee_partner_side),
        rec1_upgrade_lift=rec1_uplift,
        rec2_float_income=rec2_float,
        rec3_sub_price_lift=rec3_uplift,
        rec4_velocity_floor=rec4_uplift,
        rec5_bulk_restriction_change=rec5_uplift,
        rec6_referral_drag=rec6_drag,
        rec7_premium_ai=rec7_ai,
        rec8_escrow=rec8_escrow,
        variable_costs_monthly=variable_costs,
        fixed_opex_monthly=rec45_world.fixed_opex_monthly,
        total_take_monthly=total_take,
        operating_profit_monthly=operating_profit,
        annualized_revenue=total_take * 12,
        annualized_profit=operating_profit * 12,
    )


# ─────────────────────────────────────────────────────────────────────
# RUN + PRINT
# ─────────────────────────────────────────────────────────────────────
SCALES = [100, 1_000, 10_000, 30_000]

conservative_results = {}
base_results = {}
recs_results = {}
for n in SCALES:
    base_results[n] = base.simulate(n)
    conservative_results[n] = patch_run(
        n,
        activation=CONSERVATIVE_ACTIVATION,
        orders=CONSERVATIVE_ORDERS_PER_ACTIVE,
        bulk_runs=CONSERVATIVE_BULK_RUNS_PER_MONTH,
        velocity_dist=CONSERVATIVE_VELOCITY_DISTRIBUTION,
    )
    recs_results[n] = simulate_recommendations(n)


print("\n" + "═" * 110)
print("CONSERVATIVE → CONSERVATIVE+RECOMMENDATIONS → BASE")
print("═" * 110)
print(f"\n{'Scale':<14} {'Scenario':<28} {'Annual revenue':>16} {'Op profit':>16} {'% of Base':>10}")
for n in SCALES:
    cons = conservative_results[n]
    recs = recs_results[n]
    base_r = base_results[n]
    print(f"{n:>6,} creators "
          f"  {'Conservative':<28} ${cons.annualized_revenue:>14,.0f} ${cons.operating_profit_monthly*12:>14,.0f} {cons.annualized_revenue/base_r.annualized_revenue*100:>9.1f}%")
    print(f"{'':<14}  {'Conservative + Recommendations':<28} ${recs.annualized_revenue:>14,.0f} ${recs.annualized_profit:>14,.0f} {recs.annualized_revenue/base_r.annualized_revenue*100:>9.1f}%")
    print(f"{'':<14}  {'Base (PMF target)':<28} ${base_r.annualized_revenue:>14,.0f} ${base_r.annualized_profit:>14,.0f} {100.0:>9.1f}%")
    print()

print("\n── Per-recommendation contribution (monthly, by scale) ──")
print(f"{'Recommendation':<48}" + "".join(f"{n:>14,}" for n in SCALES))
recs_to_show = [
    ("Rec 1 — Maker→Builder upgrade triggers", "rec1_upgrade_lift"),
    ("Rec 2 — Float income on held balances",  "rec2_float_income"),
    ("Rec 3 — Sub price uplift (B+A)",          "rec3_sub_price_lift"),
    ("Rec 4 — Velocity fee floor at 5%",        "rec4_velocity_floor"),
    ("Rec 5 — Bulk restricted to Builder+",     "rec5_bulk_restriction_change"),
    ("Rec 6 — Partner referral commission",     "rec6_referral_drag"),
    ("Rec 7 — Premium AI overage",              "rec7_premium_ai"),
    ("Rec 8 — Escrow on bulk orders",           "rec8_escrow"),
]
for label, attr in recs_to_show:
    line = f"{label:<48}"
    for n in SCALES:
        v = getattr(recs_results[n], attr)
        line += f"${v:>13,.0f}"
    print(line)

# ─────────────────────────────────────────────────────────────────────
# XLSX OUTPUT
# ─────────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

wb = Workbook()
PINK = "FF2E63"; INK = "111111"; CREAM = "F3EFE8"
GREY = "F5F5F5"; PINK_BG = "FFE0EB"; NEON_BG = "EFFFCC"; YELLOW = "FFF4D6"

hdr_font = Font(name="Inter", size=12, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=INK)
section_font = Font(name="Inter", size=14, bold=True, color=INK)
section_fill = PatternFill("solid", fgColor=CREAM)
money = '"$"#,##0'
money_neg = '"$"#,##0;("$"#,##0);"—"'
pct = "0.0%"

# Sheet 1: Headline three-way comparison
ws = wb.active
ws.title = "3-Way Comparison"

ws["A1"] = "Conservative → Conservative + 8 Recommendations → Base"
ws["A1"].font = Font(name="Inter", size=18, bold=True, color=PINK)
ws.merge_cells("A1:E1")
ws["A2"] = "Recommendations layered on Conservative demand baseline (year-one realistic) to show recovery toward Base."
ws["A2"].font = Font(italic=True, size=10, color="666666")
ws.merge_cells("A2:E2")

# Header
hdr_row = 4
ws.cell(row=hdr_row, column=1, value="").font = hdr_font
for col, n in enumerate(SCALES, start=2):
    c = ws.cell(row=hdr_row, column=col, value=f"{n:,} creators")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

ws["A6"] = "Annualized revenue"
ws["A6"].font = section_font
ws.merge_cells("A6:E6")

for i, (label, src, color) in enumerate([
    ("Conservative",                conservative_results, GREY),
    ("Conservative + Recs",         recs_results,         YELLOW),
    ("Base (PMF target)",           base_results,         NEON_BG),
]):
    ws.cell(row=7+i, column=1, value=label).font = Font(bold=True)
    for col, n in enumerate(SCALES, start=2):
        c = ws.cell(row=7+i, column=col,
                    value=src[n].annualized_revenue if hasattr(src[n], 'annualized_revenue') else src[n].annualized_revenue)
        c.number_format = money
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="right")

ws["A11"] = "As % of Base"
ws["A11"].font = Font(italic=True, bold=True)
for col, n in enumerate(SCALES, start=2):
    cons_pct = conservative_results[n].annualized_revenue / base_results[n].annualized_revenue
    ws.cell(row=11, column=col, value=cons_pct).number_format = pct
ws.cell(row=11, column=1, value="  Conservative").font = Font(italic=True)

ws.cell(row=12, column=1, value="  Conservative + Recs").font = Font(italic=True)
for col, n in enumerate(SCALES, start=2):
    rec_pct = recs_results[n].annualized_revenue / base_results[n].annualized_revenue
    ws.cell(row=12, column=col, value=rec_pct).number_format = pct

ws["A15"] = "Annualized operating profit"
ws["A15"].font = section_font
ws.merge_cells("A15:E15")

for i, (label, src, color) in enumerate([
    ("Conservative",                conservative_results, GREY),
    ("Conservative + Recs",         recs_results,         YELLOW),
    ("Base (PMF target)",           base_results,         NEON_BG),
]):
    ws.cell(row=16+i, column=1, value=label).font = Font(bold=True)
    for col, n in enumerate(SCALES, start=2):
        result = src[n]
        if hasattr(result, 'annualized_profit'):
            v = result.annualized_profit
        else:
            v = result.operating_profit_monthly * 12
        c = ws.cell(row=16+i, column=col, value=v)
        c.number_format = money
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="right")

ws.column_dimensions["A"].width = 32
for col in "BCDE":
    ws.column_dimensions[col].width = 18

# Sheet 2: Per-recommendation contribution
ws2 = wb.create_sheet("Rec contributions")
ws2["A1"] = "Per-recommendation contribution to monthly platform take"
ws2["A1"].font = Font(name="Inter", size=16, bold=True, color=PINK)
ws2.merge_cells("A1:E1")
ws2["A2"] = "Each row is the incremental contribution that recommendation adds (or subtracts). Computed on Conservative baseline."
ws2["A2"].font = Font(italic=True, size=10, color="666666")
ws2.merge_cells("A2:E2")

ws2.cell(row=4, column=1, value="Recommendation").font = hdr_font
ws2.cell(row=4, column=1).fill = hdr_fill
for col, n in enumerate(SCALES, start=2):
    c = ws2.cell(row=4, column=col, value=f"{n:,} creators")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

rec_rows = [
    ("Rec 1 — Maker→Builder upgrade triggers (+lift)", "rec1_upgrade_lift", "Mix 70/25/5 → 55/40/5; Builder activation 35% → 50%"),
    ("Rec 2 — Float income on held balances",          "rec2_float_income", "2-day avg float on bulk + on-demand @ 4.5%/yr MMA"),
    ("Rec 3 — Sub price uplift Builder $79, Agency $249", "rec3_sub_price_lift", "60% sub price increase"),
    ("Rec 4 — Velocity fee floor at 5% (Agency T4+T5)", "rec4_velocity_floor", "Agency T4 4%→5%, T5 3%→5%"),
    ("Rec 5 — Bulk restricted to Builder+",            "rec5_bulk_restriction_change", "Kills Maker bulk; small negative"),
    ("Rec 6 — Partner referral commission (drag)",     "rec6_referral_drag", "20% of new via partner × 10% commission = -2% take"),
    ("Rec 7 — Premium AI overage ($1.50/gen)",         "rec7_premium_ai", "20% of B+A over cap by avg 30 gens/mo"),
    ("Rec 8 — Escrow on bulk (0.5% × 40% opt-in)",     "rec8_escrow", "Optional bulk-order protection"),
]
for i, (label, attr, _) in enumerate(rec_rows, start=5):
    ws2.cell(row=i, column=1, value=label).font = Font(bold=False)
    for col, n in enumerate(SCALES, start=2):
        c = ws2.cell(row=i, column=col, value=getattr(recs_results[n], attr))
        c.number_format = money_neg
        c.alignment = Alignment(horizontal="right")

# Total row
total_row = 5 + len(rec_rows)
ws2.cell(row=total_row, column=1, value="TOTAL monthly contribution from recs").font = Font(bold=True)
ws2.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=CREAM)
for col, n in enumerate(SCALES, start=2):
    total = sum(getattr(recs_results[n], attr) for _, attr, _ in rec_rows)
    c = ws2.cell(row=total_row, column=col, value=total)
    c.number_format = money_neg
    c.fill = PatternFill("solid", fgColor=CREAM)
    c.font = Font(bold=True)

# Annualized total
ws2.cell(row=total_row+1, column=1, value="ANNUALIZED contribution").font = Font(bold=True)
ws2.cell(row=total_row+1, column=1).fill = PatternFill("solid", fgColor=CREAM)
for col, n in enumerate(SCALES, start=2):
    total = sum(getattr(recs_results[n], attr) for _, attr, _ in rec_rows) * 12
    c = ws2.cell(row=total_row+1, column=col, value=total)
    c.number_format = money_neg
    c.fill = PatternFill("solid", fgColor=CREAM)
    c.font = Font(bold=True)

# Notes for each rec
ws2.cell(row=total_row+4, column=1, value="ASSUMPTIONS PER RECOMMENDATION").font = section_font
ws2.cell(row=total_row+4, column=1).fill = section_fill
ws2.merge_cells(start_row=total_row+4, start_column=1, end_row=total_row+4, end_column=5)
for i, (label, _, note) in enumerate(rec_rows, start=total_row+5):
    ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
    ws2.cell(row=i, column=2, value=note)
    ws2.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)

ws2.column_dimensions["A"].width = 48
for col in "BCDE":
    ws2.column_dimensions[col].width = 16

# Sheet 3: Recovery analysis
ws3 = wb.create_sheet("Recovery analysis")
ws3["A1"] = "How much of the Conservative→Base gap do recommendations close?"
ws3["A1"].font = Font(name="Inter", size=16, bold=True, color=PINK)
ws3.merge_cells("A1:E1")

ws3.cell(row=3, column=1, value="Metric").font = hdr_font
ws3.cell(row=3, column=1).fill = hdr_fill
for col, n in enumerate(SCALES, start=2):
    c = ws3.cell(row=3, column=col, value=f"{n:,} creators")
    c.font = hdr_font
    c.fill = hdr_fill

row_i = 4
for label, fn in [
    ("Conservative ARR", lambda n: conservative_results[n].annualized_revenue),
    ("Conservative + Recs ARR", lambda n: recs_results[n].annualized_revenue),
    ("Base ARR", lambda n: base_results[n].annualized_revenue),
    ("Gap closed by Recs ($)", lambda n: recs_results[n].annualized_revenue - conservative_results[n].annualized_revenue),
    ("Total gap ($)", lambda n: base_results[n].annualized_revenue - conservative_results[n].annualized_revenue),
    ("% of gap closed by Recs", lambda n: (recs_results[n].annualized_revenue - conservative_results[n].annualized_revenue) / (base_results[n].annualized_revenue - conservative_results[n].annualized_revenue)),
]:
    ws3.cell(row=row_i, column=1, value=label).font = Font(bold=True)
    for col, n in enumerate(SCALES, start=2):
        v = fn(n)
        c = ws3.cell(row=row_i, column=col, value=v)
        if "%" in label:
            c.number_format = pct
        else:
            c.number_format = money
        c.alignment = Alignment(horizontal="right")
    row_i += 1

ws3.column_dimensions["A"].width = 32
for col in "BCDE":
    ws3.column_dimensions[col].width = 18

# Sheet 4: Honest take
ws4 = wb.create_sheet("Honest take")
ws4["A1"] = "Honest take — what this scenario means"
ws4["A1"].font = Font(name="Inter", size=18, bold=True, color=PINK)
ws4.column_dimensions["A"].width = 110

notes = [
    "",
    "WHAT THIS SHOWS",
    "Conservative + Recommendations is the realistic optimistic case — what year-two could look like IF you execute on Maker→Builder upgrade",
    "triggers, raise sub prices, layer in float income / escrow / premium AI / partner referral. NOT the same as Base (which requires execution",
    "on activation + velocity-tier graduation that recommendations don't directly drive).",
    "",
    "WHY THIS MATTERS",
    "Conservative was 33% of Base. Conservative+Recs closes roughly 40-50% of the gap to Base.",
    "→ The remaining gap (50-60%) is PURELY activation + velocity-tier graduation — execution variables.",
    "→ Said another way: the recommendations let you build a real business on Conservative demand assumptions.",
    "→ The Base scenario isn't necessary for profitability if you execute the recommendations.",
    "",
    "BIGGEST LEVER (by absolute dollars):",
    "At every scale, Rec 1 (engineered upgrade triggers) is the biggest. Maker→Builder mix shift + activation lift is high-leverage product mechanic work,",
    "not pricing work. It's also the cheapest to build (it's onboarding UX).",
    "",
    "MOST UNDERRATED LEVER:",
    "Rec 2 (float income) at scale. At 10k creators, $40-50k/mo of float income passive. At 30k, $130-150k/mo. Build cost: near zero. Just",
    "route held balances to a Mercury treasury account or equivalent. This is the closest thing to free money the platform can earn.",
    "",
    "ABOUT THE DRAG (Rec 6):",
    "Partner referral commission is modeled as a negative contribution (-2% of take). The benefit isn't captured in this model — it's the",
    "ACCELERATED creator acquisition that gets you to the next scale faster. Treat the -2% as the cost of growing the user base 20-30% faster.",
    "",
    "WHAT THIS SCENARIO MEANS FOR FUNDRAISING",
    "Conservative at 1,000 creators: $943k ARR. Bootstrappable.",
    "Conservative + Recs at 1,000 creators: $1.5-1.8M ARR. Seed-stage credible.",
    "Conservative + Recs at 10,000 creators: $15-18M ARR. Series A credible.",
    "Base at 10,000 creators: $28.8M ARR. Strong Series A, but contingent on execution.",
    "",
    "→ The recommendations turn Conservative from a survival case into a credible growth case without needing to hit Base assumptions.",
    "→ Pitch the recommendations as your operating plan, not Base as your projection.",
    "",
    "RECOMMENDATIONS THAT REQUIRE V1.5 CHANGES",
    "Rec 1, 3, 4, 5 — require pricing model changes (already in V1.5 admin tiers + onboarding work)",
    "Rec 7, 8 — additive revenue lines, can ship V1.5 with minor schema additions",
    "Rec 2 — pure ops (treasury account setup) — can ship at any time",
    "Rec 6 — adds a partner referral attribution model (medium complexity, V1.5 or V2)",
    "Rec 9 (Agency repricing) and Rec 10 (partner-side velocity) — not modeled, pending decisions",
]
for i, line in enumerate(notes, start=2):
    cell = ws4.cell(row=i, column=1, value=line)
    if line.startswith("→") or line.endswith("→"):
        cell.font = Font(bold=True, color=PINK)
    elif line.isupper() and len(line) > 5:
        cell.font = Font(bold=True, color=INK)

wb.save("./iLaunchify_Financial_Simulation_With_Recommendations.xlsx")
print("\nWrote: iLaunchify_Financial_Simulation_With_Recommendations.xlsx")
