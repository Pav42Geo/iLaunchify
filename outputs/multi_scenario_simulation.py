"""
iLaunchify financial simulation — 3 scenarios side-by-side
  Base     = locked V1.5 model
  +5pp     = +5 percentage points on every velocity tier (creator AND partner side)
  +10pp    = +10 percentage points on every velocity tier

Same activation / order-velocity / distribution assumptions across all 3.
Only the fee % changes — so we isolate the pricing-lever sensitivity.
"""

import json
from copy import deepcopy
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Import the base assumptions + sim machinery from the existing model
import sys
sys.path.insert(0, "./")
from financial_simulation import (
    CREATOR_TIER_MIX, CREATOR_ACTIVATION,
    ORDERS_PER_ACTIVE_CREATOR_PER_MONTH, BULK_MIX_BY_TIER,
    AVG_BULK_RUN_SIZE, BULK_RUNS_PER_MONTH,
    SUBSCRIPTION_PRICE_MONTHLY,
    VELOCITY_FEE_PCT as BASE_VELOCITY_FEE_PCT,
    VELOCITY_DISTRIBUTION,
    ON_DEMAND_RETAIL_PRICE, ON_DEMAND_PARTNER_WHOLESALE,
    BULK_PARTNER_WHOLESALE_PER_UNIT,
    PARTNER_FEE_PCT as BASE_PARTNER_FEE_PCT,
    PARTNER_TIER_MIX,
    CREATORS_PER_PARTNER,
    STRIPE_PCT_PER_TRANSACTION, STRIPE_FLAT_PER_TRANSACTION,
    AI_PARSER_ANTHROPIC_COST_PER_PARSE,
    AI_PARSES_PER_BUILDER_PER_MONTH, AI_PARSES_PER_AGENCY_PER_MONTH,
    SHUTTERSTOCK_MONTHLY_BUDGET, fixed_opex,
    ScaleResult,
)


def bumped_fees(creator_fees, partner_fees, bump_pp):
    """Return new fee tables with +bump_pp added to every cell."""
    new_creator = {tier: {k: v + bump_pp for k, v in tiers.items()}
                   for tier, tiers in creator_fees.items()}
    new_partner = {tier: v + bump_pp for tier, v in partner_fees.items()}
    return new_creator, new_partner


def effective_creator_fee_pct(sub_tier: str, fee_table) -> float:
    return sum(fee_table[sub_tier][v_tier] * share
               for v_tier, share in VELOCITY_DISTRIBUTION[sub_tier].items())


def effective_partner_fee_pct(partner_table) -> float:
    return sum(partner_table[tier] * share
               for tier, share in PARTNER_TIER_MIX.items())


def simulate_scenario(creator_count: int, creator_fees, partner_fees) -> ScaleResult:
    r = ScaleResult(
        creator_count=creator_count,
        partner_count=max(1, round(creator_count / CREATORS_PER_PARTNER)),
    )
    registered = {tier: round(creator_count * mix) for tier, mix in CREATOR_TIER_MIX.items()}
    active = {tier: round(registered[tier] * CREATOR_ACTIVATION[tier]) for tier in registered}

    r.active_makers = active["maker"]
    r.active_builders = active["builder"]
    r.active_agencies = active["agency"]

    r.subscription_revenue_monthly = (
        registered["builder"] * SUBSCRIPTION_PRICE_MONTHLY["builder"]
        + registered["agency"] * SUBSCRIPTION_PRICE_MONTHLY["agency"]
    )

    for tier in ("maker", "builder", "agency"):
        active_count = active[tier]
        orders_per = ORDERS_PER_ACTIVE_CREATOR_PER_MONTH[tier]
        bulk_runs = BULK_RUNS_PER_MONTH[tier]
        bulk_size = AVG_BULK_RUN_SIZE[tier]
        bulk_share = BULK_MIX_BY_TIER[tier]

        total_unit_demand = active_count * orders_per
        on_demand_units = total_unit_demand * (1 - bulk_share)
        bulk_units = active_count * bulk_runs * bulk_size

        gross_od = on_demand_units * ON_DEMAND_RETAIL_PRICE
        eff_creator_fee = effective_creator_fee_pct(tier, creator_fees) / 100
        eff_partner_fee = effective_partner_fee_pct(partner_fees) / 100

        creator_fee_od = gross_od * eff_creator_fee
        partner_fee_od = on_demand_units * ON_DEMAND_PARTNER_WHOLESALE * eff_partner_fee

        bulk_gross_partner_cost = bulk_units * BULK_PARTNER_WHOLESALE_PER_UNIT
        creator_fee_bulk = bulk_gross_partner_cost * eff_creator_fee
        partner_fee_bulk = bulk_gross_partner_cost * eff_partner_fee

        r.on_demand_units_monthly += round(on_demand_units)
        r.on_demand_gross_revenue_monthly += gross_od
        r.on_demand_platform_fee_creator_side += creator_fee_od
        r.on_demand_platform_fee_partner_side += partner_fee_od
        r.bulk_units_monthly += round(bulk_units)
        r.bulk_gross_revenue_monthly += bulk_gross_partner_cost
        r.bulk_platform_fee_creator_side += creator_fee_bulk
        r.bulk_platform_fee_partner_side += partner_fee_bulk

    total_transactions = r.on_demand_units_monthly + (r.bulk_units_monthly // 500)
    r.stripe_costs_monthly = (
        (r.on_demand_gross_revenue_monthly + r.bulk_gross_revenue_monthly) * STRIPE_PCT_PER_TRANSACTION
        + total_transactions * STRIPE_FLAT_PER_TRANSACTION
    )
    r.ai_parser_costs_monthly = (
        round(creator_count * CREATOR_TIER_MIX["builder"] * CREATOR_ACTIVATION["builder"])
            * AI_PARSES_PER_BUILDER_PER_MONTH * AI_PARSER_ANTHROPIC_COST_PER_PARSE
        + round(creator_count * CREATOR_TIER_MIX["agency"] * CREATOR_ACTIVATION["agency"])
            * AI_PARSES_PER_AGENCY_PER_MONTH * AI_PARSER_ANTHROPIC_COST_PER_PARSE
    )
    r.asset_library_costs_monthly = SHUTTERSTOCK_MONTHLY_BUDGET
    r.fixed_opex_monthly = fixed_opex(creator_count) / 12

    r.total_platform_take_monthly = (
        r.subscription_revenue_monthly
        + r.on_demand_platform_fee_creator_side
        + r.on_demand_platform_fee_partner_side
        + r.bulk_platform_fee_creator_side
        + r.bulk_platform_fee_partner_side
    )
    r.total_variable_costs_monthly = (
        r.stripe_costs_monthly + r.ai_parser_costs_monthly + r.asset_library_costs_monthly
    )
    r.net_revenue_monthly = r.total_platform_take_monthly - r.total_variable_costs_monthly
    r.operating_profit_monthly = r.net_revenue_monthly - r.fixed_opex_monthly
    r.annualized_revenue = r.total_platform_take_monthly * 12
    r.annualized_profit = r.operating_profit_monthly * 12
    return r


# ─────────────────────────────────────────────────────────────────────
# 3 scenarios × 4 scales
# ─────────────────────────────────────────────────────────────────────
SCENARIOS = [
    ("Base (locked)", BASE_VELOCITY_FEE_PCT, BASE_PARTNER_FEE_PCT),
    ("+5pp uplift", *bumped_fees(BASE_VELOCITY_FEE_PCT, BASE_PARTNER_FEE_PCT, 5.0)),
    ("+10pp uplift", *bumped_fees(BASE_VELOCITY_FEE_PCT, BASE_PARTNER_FEE_PCT, 10.0)),
]
SCALES = [100, 1_000, 10_000, 30_000]

results = {}
for name, creator_fees, partner_fees in SCENARIOS:
    results[name] = {n: simulate_scenario(n, creator_fees, partner_fees) for n in SCALES}

# Print summary
print("\n" + "═" * 110)
print("iLaunchify financial simulation — 3 scenarios")
print("═" * 110)
for name, creator_fees, partner_fees in SCENARIOS:
    m_eff = effective_creator_fee_pct("maker", creator_fees)
    b_eff = effective_creator_fee_pct("builder", creator_fees)
    a_eff = effective_creator_fee_pct("agency", creator_fees)
    p_eff = effective_partner_fee_pct(partner_fees)
    print(f"\n── {name} ──")
    print(f"  Effective Maker/Builder/Agency: {m_eff:.2f}% / {b_eff:.2f}% / {a_eff:.2f}%")
    print(f"  Effective partner-side: {p_eff:.2f}%")
    for n in SCALES:
        r = results[name][n]
        print(f"  {n:>6,} creators → Annualized rev ${r.annualized_revenue:>14,.0f}  · Op profit ${r.annualized_profit:>14,.0f}")

# ─────────────────────────────────────────────────────────────────────
# XLSX output
# ─────────────────────────────────────────────────────────────────────
wb = Workbook()
PINK = "FF2E63"
INK_900 = "111111"
CREAM = "F3EFE8"
LIGHT_PINK = "FFE0EB"
LIGHT_NEON = "EFFFCC"

hdr_font = Font(name="Inter", size=12, bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor=INK_900)
section_font = Font(name="Inter", size=14, bold=True, color=INK_900)
section_fill = PatternFill("solid", fgColor=CREAM)
money_format = '"$"#,##0'
pct_format = '0.00"%"'

# ─── Sheet 1: Headline comparison ───
ws = wb.active
ws.title = "Headline comparison"
ws["A1"] = "iLaunchify financial simulation — 3 scenarios × 4 scales"
ws["A1"].font = Font(name="Inter", size=18, bold=True, color=PINK)
ws.merge_cells("A1:E1")
ws["A2"] = "Same activation + velocity + order-volume assumptions across scenarios. ONLY fee % differs."
ws["A2"].font = Font(italic=True, size=10, color="666666")
ws.merge_cells("A2:E2")

# Effective fees summary
ws["A4"] = "Effective fee % per scenario (volume-weighted)"
ws["A4"].font = section_font
ws.merge_cells("A4:E4")

ws.cell(row=5, column=1, value="Scenario").font = hdr_font
ws.cell(row=5, column=1).fill = hdr_fill
for col, label in enumerate(["Maker", "Builder", "Agency", "Partner avg"], start=2):
    c = ws.cell(row=5, column=col, value=label)
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

for row_i, (name, creator_fees, partner_fees) in enumerate(SCENARIOS, start=6):
    ws.cell(row=row_i, column=1, value=name).font = Font(bold=True)
    ws.cell(row=row_i, column=2, value=effective_creator_fee_pct("maker", creator_fees)).number_format = pct_format
    ws.cell(row=row_i, column=3, value=effective_creator_fee_pct("builder", creator_fees)).number_format = pct_format
    ws.cell(row=row_i, column=4, value=effective_creator_fee_pct("agency", creator_fees)).number_format = pct_format
    ws.cell(row=row_i, column=5, value=effective_partner_fee_pct(partner_fees)).number_format = pct_format

# Annualized revenue comparison
ws["A10"] = "Annualized platform revenue (ARR-equivalent)"
ws["A10"].font = section_font
ws.merge_cells("A10:E10")

ws.cell(row=11, column=1, value="Scenario").font = hdr_font
ws.cell(row=11, column=1).fill = hdr_fill
for col, n in enumerate(SCALES, start=2):
    c = ws.cell(row=11, column=col, value=f"{n:,} creators")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

for row_i, (name, *_) in enumerate(SCENARIOS, start=12):
    ws.cell(row=row_i, column=1, value=name).font = Font(bold=True)
    for col, n in enumerate(SCALES, start=2):
        cell = ws.cell(row=row_i, column=col, value=results[name][n].annualized_revenue)
        cell.number_format = money_format
        cell.alignment = Alignment(horizontal="right")
        if name.startswith("+5"):
            cell.fill = PatternFill("solid", fgColor=LIGHT_PINK)
        elif name.startswith("+10"):
            cell.fill = PatternFill("solid", fgColor=LIGHT_NEON)

# Annualized operating profit comparison
ws["A16"] = "Annualized operating profit"
ws["A16"].font = section_font
ws.merge_cells("A16:E16")

ws.cell(row=17, column=1, value="Scenario").font = hdr_font
ws.cell(row=17, column=1).fill = hdr_fill
for col, n in enumerate(SCALES, start=2):
    c = ws.cell(row=17, column=col, value=f"{n:,} creators")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

for row_i, (name, *_) in enumerate(SCENARIOS, start=18):
    ws.cell(row=row_i, column=1, value=name).font = Font(bold=True)
    for col, n in enumerate(SCALES, start=2):
        cell = ws.cell(row=row_i, column=col, value=results[name][n].annualized_profit)
        cell.number_format = money_format
        cell.alignment = Alignment(horizontal="right")
        if name.startswith("+5"):
            cell.fill = PatternFill("solid", fgColor=LIGHT_PINK)
        elif name.startswith("+10"):
            cell.fill = PatternFill("solid", fgColor=LIGHT_NEON)

# Revenue uplift comparison (vs base)
ws["A22"] = "Revenue uplift vs Base scenario (additional ARR)"
ws["A22"].font = section_font
ws.merge_cells("A22:E22")

ws.cell(row=23, column=1, value="Scenario").font = hdr_font
ws.cell(row=23, column=1).fill = hdr_fill
for col, n in enumerate(SCALES, start=2):
    c = ws.cell(row=23, column=col, value=f"{n:,} creators")
    c.font = hdr_font
    c.fill = hdr_fill
    c.alignment = Alignment(horizontal="center")

base_rev_by_scale = {n: results["Base (locked)"][n].annualized_revenue for n in SCALES}
for row_i, (name, *_) in enumerate(SCENARIOS, start=24):
    ws.cell(row=row_i, column=1, value=name).font = Font(bold=True)
    for col, n in enumerate(SCALES, start=2):
        delta = results[name][n].annualized_revenue - base_rev_by_scale[n]
        cell = ws.cell(row=row_i, column=col, value=delta)
        cell.number_format = '"+$"#,##0;"-$"#,##0;"—"'
        cell.alignment = Alignment(horizontal="right")
        if name.startswith("+5"):
            cell.fill = PatternFill("solid", fgColor=LIGHT_PINK)
        elif name.startswith("+10"):
            cell.fill = PatternFill("solid", fgColor=LIGHT_NEON)

ws.column_dimensions["A"].width = 22
for col in "BCDE":
    ws.column_dimensions[col].width = 20

# ─── Per-scenario detail sheets ───
def write_detail(ws, name, scenario_results):
    ws["A1"] = f"Scenario: {name}"
    ws["A1"].font = Font(name="Inter", size=16, bold=True, color=PINK)
    ws.merge_cells("A1:E1")

    header = ["Metric"] + [f"{n:,} creators" for n in SCALES]
    for col, h in enumerate(header, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")

    def row(label, key, fmt=money_format, bold=False):
        idx = ws.max_row + 1
        ws.cell(row=idx, column=1, value=label).font = Font(bold=bold)
        for col, n in enumerate(SCALES, start=2):
            cell = ws.cell(row=idx, column=col, value=getattr(scenario_results[n], key))
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="right")
        return idx

    def section(label):
        idx = ws.max_row + 2
        c = ws.cell(row=idx, column=1, value=label)
        c.font = section_font
        c.fill = section_fill
        return idx

    section("USER COUNTS")
    row("Partners", "partner_count", "#,##0")
    row("Active Makers", "active_makers", "#,##0")
    row("Active Builders", "active_builders", "#,##0")
    row("Active Agencies", "active_agencies", "#,##0")

    section("SUBSCRIPTION (monthly)")
    row("Subscription revenue", "subscription_revenue_monthly")

    section("ON-DEMAND (monthly)")
    row("On-demand units", "on_demand_units_monthly", "#,##0")
    row("Gross retail", "on_demand_gross_revenue_monthly")
    row("Creator-side fee", "on_demand_platform_fee_creator_side")
    row("Partner-side fee", "on_demand_platform_fee_partner_side")

    section("BULK (monthly)")
    row("Bulk units", "bulk_units_monthly", "#,##0")
    row("Partner cost flowing", "bulk_gross_revenue_monthly")
    row("Creator-side fee", "bulk_platform_fee_creator_side")
    row("Partner-side fee", "bulk_platform_fee_partner_side")

    section("COSTS (monthly)")
    row("Stripe Connect", "stripe_costs_monthly")
    row("AI parser (Anthropic)", "ai_parser_costs_monthly")
    row("Asset library", "asset_library_costs_monthly")
    row("Fixed opex", "fixed_opex_monthly")

    section("BOTTOM LINE")
    row("Total platform take (mo)", "total_platform_take_monthly", bold=True)
    row("Operating profit (mo)", "operating_profit_monthly", bold=True)
    row("Annualized revenue", "annualized_revenue", bold=True)
    row("Annualized op profit", "annualized_profit", bold=True)

    # Highlight bottom 2 rows
    for r_idx in (ws.max_row - 1, ws.max_row):
        for col_idx in range(1, 6):
            ws.cell(row=r_idx, column=col_idx).fill = PatternFill("solid", fgColor=CREAM)
            ws.cell(row=r_idx, column=col_idx).font = Font(bold=True, size=12)

    ws.column_dimensions["A"].width = 32
    for col in "BCDE":
        ws.column_dimensions[col].width = 18


for name, *_ in SCENARIOS:
    detail_ws = wb.create_sheet(name)
    write_detail(detail_ws, name, results[name])

# ─── Honest-take sheet ───
ws_h = wb.create_sheet("Honest take")
ws_h["A1"] = "Honest take — read before quoting any of these numbers"
ws_h["A1"].font = Font(name="Inter", size=16, bold=True, color=PINK)
ws_h.column_dimensions["A"].width = 110

notes = [
    "",
    "THE THREE SCENARIOS",
    "Base (locked):  the V1.5 locked fee structure. Effective Maker 14.76% / Builder 8.92% / Agency 5.62% / Partner 4.47%.",
    "+5pp uplift:    every velocity-tier cell + every partner-tier cell rises by 5 percentage points absolute.",
    "                Effective Maker 19.76% / Builder 13.92% / Agency 10.62% / Partner 9.47%.",
    "+10pp uplift:   every cell rises by 10pp absolute.",
    "                Effective Maker 24.76% / Builder 18.92% / Agency 15.62% / Partner 14.47%.",
    "",
    "MARKET CONTEXT — what comparable platforms charge",
    "Shopify Payments take rate (creator-facing): 2.9% + $0.30",
    "Printful base markup (white-label creator-facing): ~30% of partner cost (built into product price, not labeled as fee)",
    "Printify base markup: ~30-40% of partner cost",
    "Supliful subscription fee: $49/mo + product wholesale (no separate platform %)",
    "iLaunchify base (Maker on-demand): 15% of retail = effectively 50% of partner wholesale at $30 retail / $8 wholesale",
    "",
    "TRANSLATING TO PARTNER-WHOLESALE COMPARISON (apples to apples)",
    "Maker Base 15% of $30 retail = $4.50 fee, which is 56% of the $8 partner wholesale.",
    "Maker +5pp = 20% of retail = $6.00 fee = 75% of partner wholesale.",
    "Maker +10pp = 25% of retail = $7.50 fee = 94% of partner wholesale.",
    "",
    "→ +10pp on Maker tier is operationally untenable. The fee almost equals the partner cost.",
    "→ +5pp on Maker tier is borderline acceptable for free tier creators; +5pp on Builder/Agency is competitive.",
    "",
    "PARTNER-SIDE STRESS TEST",
    "Partners typically operate on 15-30% margin on their wholesale price.",
    "Base 5% Verified is reasonable — eats ~17-33% of partner margin.",
    "+5pp Verified at 10% — eats 33-67% of partner margin. Partners will negotiate or churn.",
    "+10pp Verified at 15% — eats 50-100% of partner margin. Most partners walk.",
    "",
    "→ Don't model partner-side bumps as freely tuneable. They're constrained by partner unit economics.",
    "→ A more realistic 'aggressive' scenario keeps partner side at Base and only bumps creator side.",
    "",
    "WHAT YOU CAN ACTUALLY DO WITH THESE SCENARIOS",
    "1. Identify pricing-power ceiling. The Base scenario is competitive; +5pp creator-side is the realistic uplift before churn pressure.",
    "2. Stress-test fundraising narrative. Run the spreadsheet at 50% activation to see realistic year-one numbers.",
    "3. Negotiate partner economics. Show partners the BASE scenario; never the +pp scenarios.",
    "4. Plan compensation. At 1,000 creators on the +5pp scenario, ARR is $4.5M-ish — you can hire aggressively.",
    "",
    "WHAT YOU SHOULD NOT DO",
    "- Don't show the +10pp scenario externally (it implies pricing power you haven't proven).",
    "- Don't price the platform at +10pp from launch (Maker churn will be catastrophic).",
    "- Don't assume partners will absorb the +5pp partner-side bump silently. Run the +5pp creator-only variant for a realistic upside case.",
]
for i, line in enumerate(notes, start=2):
    cell = ws_h.cell(row=i, column=1, value=line)
    if line.endswith("→") or line.startswith("→"):
        cell.font = Font(bold=True, color=PINK)
    elif line.isupper() and len(line) > 5:
        cell.font = Font(bold=True, color=INK_900)

wb.save("./iLaunchify_Financial_Simulation_3Scenarios.xlsx")
print("\nWrote: iLaunchify_Financial_Simulation_3Scenarios.xlsx")
